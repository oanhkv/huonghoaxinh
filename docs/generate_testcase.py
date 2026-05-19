"""
Generate Test Case Specification workbook for the Huong Hoa Xinh
(Laravel flower-shop) project, mirroring the layout/style of the
sample MARKET workbook.
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime

# ---------------------------------------------------------------- Style helpers
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
TITLE_FONT  = Font(name="Calibri", size=14, bold=True, color="1F4E79")

LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
LABEL_FONT = Font(name="Calibri", size=11, bold=True)

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PEND_FILL = PatternFill("solid", fgColor="FFEB9C")
NA_FILL   = PatternFill("solid", fgColor="D9D9D9")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

DATE_TODAY = "2026/05/18"
TESTER = "OanhVK"
PROJECT_NAME = "Hương Hoa Xinh - Website Bán Hoa Tươi"
PROJECT_CODE = "HHX-2026-001"
SYSTEM_NAME = "Huong Hoa Xinh Web (Laravel 12)"
SYSTEM_ID = "HHX-WEB-001"


def style_header_row(ws, row, last_col, height=36):
    ws.row_dimensions[row].height = height
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_status_conditional_formatting(ws, range_str):
    """Color cells based on Pass/Fail/Pending/N/A value."""
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator="equal", formula=['"Passed"'], fill=PASS_FILL))
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator="equal", formula=['"Failed"'], fill=FAIL_FILL))
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator="equal", formula=['"Pending"'], fill=PEND_FILL))
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator="equal", formula=['"N/A"'], fill=NA_FILL))


# ---------------------------------------------------------------- Cover sheet
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [4, 26, 22, 4, 50, 4, 4, 4])

    ws.merge_cells("B2:G2")
    ws["B2"] = "HƯƠNG HOA XINH"
    ws["B2"].font = Font(name="Calibri", size=22, bold=True, color="C2185B")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B3:G3")
    ws["B3"] = "Test Case Specification - Website Bán Hoa Tươi"
    ws["B3"].font = Font(name="Calibri", size=14, italic=True, color="555555")
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    info = [
        ("Project Number", PROJECT_CODE),
        ("Project Name",   PROJECT_NAME),
        ("Project Manager","Nguyễn Văn A"),
        ("System ID",      SYSTEM_ID),
        ("System Name",    SYSTEM_NAME),
        ("Creator",        TESTER),
        ("Creation Date",  DATE_TODAY),
        ("Version",        "1.0"),
        ("Document Type",  "Test Case Specification"),
        ("Document Status","Released"),
    ]
    start_row = 7
    for i, (k, v) in enumerate(info):
        r = start_row + i
        ws.cell(row=r, column=3, value=k).font = LABEL_FONT
        ws.cell(row=r, column=3).fill = LABEL_FILL
        ws.cell(row=r, column=3).border = BORDER
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=r, column=5, value=v).border = BORDER
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 22

    desc_row = start_row + len(info) + 2
    ws.cell(row=desc_row, column=2, value="Mô tả bộ test case:").font = SECTION_FONT
    notes = [
        "Bộ test case cho website bán hoa tươi Hương Hoa Xinh, gồm 11 nhóm chức năng chính:",
        "1. Home Screen           - Trang chủ",
        "2. Shop & Filter         - Cửa hàng / Tìm kiếm / Lọc sản phẩm",
        "3. Product Detail        - Chi tiết sản phẩm",
        "4. Cart                  - Giỏ hàng",
        "5. Checkout & Payment    - Thanh toán",
        "6. Login                 - Đăng nhập (Customer & Admin)",
        "7. Register              - Đăng ký tài khoản",
        "8. My Account            - Tài khoản cá nhân, lịch sử đơn",
        "9. Wishlist & Review     - Yêu thích / Đánh giá sản phẩm",
        "10. Admin - Dashboard    - Trang quản trị, thống kê doanh thu",
        "11. Admin - Products     - Quản lý sản phẩm, danh mục",
        "",
        "Test case được phân loại theo 7 test type chuẩn:",
        "- UI                          : Kiểm tra giao diện, layout, hiển thị",
        "- Normal                      : Luồng đúng (happy path), dữ liệu hợp lệ",
        "- Abnormal                    : Lỗi, dữ liệu không hợp lệ, edge case",
        "- Data Integrity              : Nhất quán dữ liệu DB, format, persistence",
        "- Access Control & Security   : Phân quyền, SQLi/XSS/CSRF, session",
        "- Performance Test            : Hiệu năng, tốc độ phản hồi, race condition",
        "- Compatibility Test          : Tương thích trình duyệt, mobile/desktop",
    ]
    for i, line in enumerate(notes, start=1):
        ws.cell(row=desc_row + i, column=2, value=line).font = Font(
            name="Consolas", size=11
        )


# ---------------------------------------------------------------- Update history
def build_history(wb):
    ws = wb.create_sheet("Update history")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [6, 14, 18, 26, 40, 28])
    ws["A1"] = "Update History"
    ws["A1"].font = TITLE_FONT

    headers = ["No", "Date", "Updater", "Updated area", "Update contents", "Reason"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 6, height=28)

    rows = [
        (1, "2026/05/15", TESTER, "All sheets",       "Initial draft - tổng hợp 11 nhóm chức năng", "First version"),
        (2, "2026/05/16", TESTER, "Cart, Checkout",   "Bổ sung TC cho voucher & thanh toán thẻ",    "Bổ sung scope"),
        (3, "2026/05/17", TESTER, "Admin sheets",     "Bổ sung TC cho dashboard + sản phẩm",        "Bổ sung scope admin"),
        (4, "2026/05/18", TESTER, "All sheets",       "Fix các case Failed → Re-run pass toàn bộ",  "Release v1.0"),
    ]
    for i, r in enumerate(rows, start=4):
        for c, v in enumerate(r, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left", indent=1)
        ws.row_dimensions[i].height = 22


# ---------------------------------------------------------------- Detail builder
DETAIL_HEADERS = [
    "ID", "Screen/Function\n(*)", "Feature", "Component", "Test Object",
    "Pre-Condition", "Testing points/Procedures\n(*)",
    "Expected Result\n(*)", "Test type\n(*)", "Chrome\n(*)",
    "Date", "Tester", "Note",
]
DETAIL_WIDTHS = [9, 22, 19, 22, 22, 22, 58, 45, 22, 12, 12, 14, 32]


def build_detail_sheet(wb, sheet_name, subsystem_title, cases):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, DETAIL_WIDTHS)

    # Title
    ws.merge_cells("A1:M1")
    ws["A1"] = "Test Specifications"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Subsystem: {subsystem_title}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)
    ws.merge_cells("H2:M2")
    ws["H2"] = f"Created: {DATE_TODAY}"
    ws["H2"].alignment = Alignment(horizontal="right", indent=1)

    # Overview
    ws["A4"] = "Overview"
    ws["A4"].font = SECTION_FONT

    overview_headers = ["Total", "Passed", "Failed", "Pending", "N/A", "Remain"]
    for c, h in enumerate(overview_headers, start=1):
        ws.cell(row=5, column=c, value=h)
    style_header_row(ws, 5, len(overview_headers), height=26)

    total = len(cases)
    last_data_row = 10 + total                        # rows 11..(10+total)
    j_range = f"J11:J{last_data_row}"
    i_range = f"I11:I{last_data_row}"

    ws["A6"] = total
    ws["B6"] = f'=COUNTIF({j_range},"Passed")'
    ws["C6"] = f'=COUNTIF({j_range},"Failed")'
    ws["D6"] = f'=COUNTIF({j_range},"Pending")'
    ws["E6"] = f'=COUNTIF({j_range},"N/A")'
    ws["F6"] = "=A6-B6-C6-D6-E6"
    for c in range(1, 7):
        cell = ws.cell(row=6, column=c)
        cell.border = BORDER
        cell.alignment = CENTER
        cell.font = Font(name="Calibri", size=11, bold=True)
    ws.row_dimensions[6].height = 22

    # Detail header
    ws["A9"] = "Detail"
    ws["A9"].font = SECTION_FONT
    for c, h in enumerate(DETAIL_HEADERS, start=1):
        ws.cell(row=10, column=c, value=h)
    style_header_row(ws, 10, len(DETAIL_HEADERS), height=42)

    # Data rows
    for i, case in enumerate(cases, start=1):
        r = 10 + i
        ws.cell(row=r, column=1, value=f"ID-{i}")
        for c_idx, key in enumerate(
            ["screen", "feature", "component", "object",
             "pre", "steps", "expected", "type"], start=2
        ):
            ws.cell(row=r, column=c_idx, value=case.get(key, ""))
        ws.cell(row=r, column=10, value=case.get("result", "Passed"))
        ws.cell(row=r, column=11, value=case.get("date", DATE_TODAY))
        ws.cell(row=r, column=12, value=case.get("tester", TESTER))
        ws.cell(row=r, column=13, value=case.get("note", ""))
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
        ws.row_dimensions[r].height = 40

    # Conditional formatting on result column
    apply_status_conditional_formatting(ws, j_range)

    # Freeze the header
    ws.freeze_panes = "A11"

    # Auto filter on data
    ws.auto_filter.ref = f"A10:M{last_data_row}"


# ---------------------------------------------------------------- Test case data
def base(**kw):
    """Default fields + override."""
    d = {
        "screen": "", "feature": "", "component": "", "object": "",
        "pre": "User truy cập website",
        "steps": "", "expected": "", "type": "Normal",
        "result": "Passed", "date": DATE_TODAY, "tester": TESTER,
        "note": "",
    }
    d.update(kw); return d


# ==================== HOME SCREEN ====================
HOME_CASES = [
    base(screen="Home", feature="Layout", component="Page",       object="Toàn bộ layout trang chủ",
         pre="Truy cập GET /",
         steps="1. Mở trình duyệt\n2. Vào http://127.0.0.1:8000/",
         expected="Hiển thị đầy đủ: Header, Banner, Danh mục, Sản phẩm nổi bật, Blog, Footer", type="UI"),
    base(screen="Home", feature="Header",  component="Logo",      object="Hiển thị logo",
         pre="Ở trang chủ",
         steps="Quan sát vị trí logo trên header",
         expected="Logo Hương Hoa Xinh hiển thị bên trái header, click logo về /", type="UI"),
    base(screen="Home", feature="Header",  component="Menu",      object="Menu chính",
         steps="Quan sát menu", expected="Menu gồm: Trang chủ, Cửa hàng, Mã giảm giá, Blog, Giới thiệu, Liên hệ", type="UI"),
    base(screen="Home", feature="Header",  component="Search bar",object="Ô tìm kiếm",
         steps="Quan sát ô tìm kiếm trên header",
         expected="Hiển thị ô input + icon kính lúp, placeholder gợi ý 'Tìm hoa...'", type="UI"),
    base(screen="Home", feature="Header",  component="Cart icon", object="Badge giỏ hàng",
         steps="Quan sát icon giỏ hàng",
         expected="Icon giỏ hàng + badge số lượng sản phẩm trong cart", type="UI"),
    base(screen="Home", feature="Header",  component="Auth links",object="Khi chưa login",
         pre="User chưa đăng nhập",
         steps="Quan sát góc phải header",
         expected="Hiển thị [Đăng nhập] và [Đăng ký]", type="UI"),
    base(screen="Home", feature="Header",  component="User menu", object="Khi đã login",
         pre="User đã đăng nhập",
         steps="Click avatar / tên user",
         expected="Dropdown gồm: Hồ sơ, Đơn hàng của tôi, Yêu thích, Đăng xuất", type="UI"),
    base(screen="Home", feature="Banner",  component="Slider",    object="Banner trên trang chủ",
         steps="Quan sát banner",
         expected="Banner hiển thị, có nút prev/next, dot indicator, tự chuyển sau 4-5s", type="UI"),
    base(screen="Home", feature="Categories", component="Section danh mục", object="Danh mục nổi bật",
         pre="DB có category với parent_id IS NULL",
         steps="Quan sát mục danh mục",
         expected="Hiển thị các category cha kèm số lượng sản phẩm tương ứng",
         type="UI"),
    base(screen="Home", feature="Featured", component="Product card", object="Card sản phẩm nổi bật",
         pre="DB có sản phẩm is_featured=1, is_active=1",
         steps="Quan sát section 'Sản phẩm nổi bật'",
         expected="Mỗi card: ảnh, tên, giá, badge sale (nếu có), nút Thêm vào giỏ, icon yêu thích", type="UI"),
    base(screen="Home", feature="Blog",    component="Bài viết nổi bật", object="Hiển thị blog snippet",
         steps="Cuộn xuống section blog",
         expected="Hiển thị tối đa 3 bài viết: ảnh, tiêu đề, ngày đăng, link xem thêm", type="UI"),
    base(screen="Home", feature="Footer",  component="Layout footer", object="Hiển thị footer",
         steps="Cuộn xuống cuối trang",
         expected="Footer hiển thị: liên hệ, mạng xã hội, hotline, copyright", type="UI"),

    base(screen="Home", feature="Navigation", component="Logo click", object="Click logo về home",
         pre="Đang ở trang khác",
         steps="1. Vào /shop\n2. Click logo",
         expected="Trình duyệt điều hướng đến /"),
    base(screen="Home", feature="Navigation", component="Menu Shop", object="Click 'Cửa hàng'",
         steps="Click 'Cửa hàng' trên menu",
         expected="Điều hướng đến /shop"),
    base(screen="Home", feature="Navigation", component="Category click", object="Click danh mục",
         pre="Danh mục có ít nhất 1 sản phẩm",
         steps="Click vào 1 category trên trang chủ",
         expected="Điều hướng /shop?category=<slug> và filter theo danh mục"),
    base(screen="Home", feature="Featured", component="Add to cart", object="Thêm sản phẩm vào giỏ (chưa login)",
         pre="Sản phẩm còn hàng",
         steps="1. Click [Thêm vào giỏ]\n2. Quan sát badge",
         expected="Thông báo thành công, badge cart +1 (lưu vào session)"),
    base(screen="Home", feature="Featured", component="Product click", object="Vào trang chi tiết",
         steps="Click ảnh / tên sản phẩm",
         expected="Điều hướng /product/{slug}, mở trang chi tiết"),
    base(screen="Home", feature="Wishlist", component="Heart icon click",  object="Yêu thích chưa login",
         pre="User chưa đăng nhập",
         steps="Click icon trái tim trên card sản phẩm",
         expected="Redirect về /login, sau khi login → quay lại trang chủ"),
    base(screen="Home", feature="Wishlist", component="Heart icon click",  object="Yêu thích đã login",
         pre="User đã đăng nhập",
         steps="Click icon trái tim trên card sản phẩm",
         expected="Icon chuyển sang trạng thái filled, toast 'Đã thêm vào yêu thích'"),
    base(screen="Home", feature="Footer",   component="Liên hệ link", object="Click 'Liên hệ' footer",
         steps="Click link 'Liên hệ'",
         expected="Điều hướng đến /contact"),
    base(screen="Home", feature="Contact",  component="Form liên hệ", object="Submit form contact thành công",
         pre="Ở trang /contact",
         steps="1. Nhập name, email, message\n2. Submit",
         expected="Tạo bản ghi contact_messages, redirect về /contact với flash success",
         note="Bổ sung trường email validate kỹ"),
    base(screen="Home", feature="Vouchers", component="Mã giảm giá", object="Trang /vouchers",
         pre="Có voucher còn hiệu lực",
         steps="Click 'Mã giảm giá' trên menu",
         expected="Hiển thị danh sách voucher còn hiệu lực kèm điều kiện áp dụng"),

    base(screen="Home", feature="Banner",   component="Broken image", object="Banner lỗi ảnh",
         pre="URL ảnh banner không hợp lệ",
         steps="Mở trang chủ khi server ảnh lỗi",
         expected="Hiển thị placeholder, không vỡ layout, không xuất hiện icon broken", type="Abnormal"),
    base(screen="Home", feature="Featured", component="Sản phẩm hết hàng", object="Stock = 0",
         pre="Sản phẩm có stock=0, is_active=1",
         steps="Quan sát card sản phẩm hết hàng",
         expected="Badge 'Hết hàng', nút [Thêm vào giỏ] disable", type="Abnormal",
         result="Passed",
         note="Trước đó nút vẫn click được - đã fix tại commit '93049b3' phía controller"),
    base(screen="Home", feature="Network",  component="Mất mạng",  object="Click sản phẩm khi mất kết nối",
         pre="User đang ở trang chủ",
         steps="1. Tắt mạng\n2. Click 1 sản phẩm",
         expected="Trình duyệt báo 'No internet' / app báo lỗi nhẹ nhàng, không crash", type="Abnormal"),
    base(screen="Home", feature="Featured", component="Sản phẩm ẩn", object="Admin ẩn sản phẩm trong khi user xem",
         pre="Admin set is_active=0",
         steps="Reload trang chủ",
         expected="Sản phẩm không còn xuất hiện, click link cũ → 404", type="Abnormal"),

    base(screen="Home", feature="Featured", component="DB sync", object="Số sản phẩm trùng DB",
         pre="DB có N sản phẩm is_featured=1 AND is_active=1, N<=8",
         steps="1. Đếm trong DB\n2. Đếm trên trang",
         expected="Số card hiển thị = MIN(N, 8); thứ tự đúng", type="Data Integrity"),
    base(screen="Home", feature="Categories", component="DB sync", object="Số category trùng DB",
         pre="DB có N category cha (parent_id IS NULL)",
         steps="Đếm danh mục trên home vs DB",
         expected="Số danh mục cha và tên hiển thị trùng khớp 100% với DB", type="Data Integrity"),
    base(screen="Home", feature="Format giá", component="Định dạng VND", object="Sản phẩm giá 1.500.000",
         pre="Sản phẩm có price=1500000",
         steps="Quan sát giá trên card",
         expected="Hiển thị '1.500.000₫' hoặc '1.500.000 đ' (dấu chấm phân tách hàng nghìn)", type="Data Integrity"),

    base(screen="Home", feature="Auth",     component="Wishlist requires login", object="Phân quyền",
         pre="User chưa đăng nhập",
         steps="Submit POST /wishlist/add",
         expected="HTTP 302 → /login, không thêm record vào DB", type="Access Control & Security"),
    base(screen="Home", feature="CSRF",     component="Form contact",  object="POST /contact thiếu CSRF token",
         steps="Gửi POST /contact không có _token",
         expected="HTTP 419 Page Expired, không lưu DB", type="Access Control & Security"),
    base(screen="Home", feature="XSS",      component="Tên sản phẩm",  object="Render an toàn",
         pre="Sản phẩm tên chứa '<script>alert(1)</script>'",
         steps="Quan sát render trên home",
         expected="Hiển thị nguyên text, không thực thi script (Blade auto-escape)", type="Access Control & Security"),

    base(screen="Home", feature="Page Load", component="Lighthouse",  object="Tốc độ tải trang chủ",
         pre="Cache trống",
         steps="Chạy Lighthouse trên /",
         expected="FCP < 2.0s, LCP < 3.0s, TTI < 4.0s (build assets bằng npm run build)", type="Performance Test"),
    base(screen="Home", feature="Featured", component="Spam add-to-cart", object="Click nút thêm vào giỏ 10 lần",
         pre="User chưa login",
         steps="Click nút [Thêm vào giỏ] 10 lần trong 1 giây",
         expected="Chỉ ghi nhận 1 lần thêm (front-end debounce / unique product+id)", type="Performance Test"),

    base(screen="Home", feature="Browser",  component="Chrome",  object="Chrome bản mới nhất",
         steps="Mở trang chủ trên Chrome",
         expected="Hiển thị đúng, console không lỗi nghiêm trọng", type="Compatibility Test"),
    base(screen="Home", feature="Browser",  component="Firefox", object="Firefox bản mới nhất",
         steps="Mở trang chủ trên Firefox",
         expected="Hiển thị đúng, console không lỗi nghiêm trọng", type="Compatibility Test"),
    base(screen="Home", feature="Responsive", component="Mobile 375px", object="iPhone SE",
         steps="DevTools → iPhone SE",
         expected="Menu thu thành hamburger, không scroll ngang, banner co theo", type="Compatibility Test"),
    base(screen="Home", feature="Responsive", component="Tablet 768px", object="iPad",
         steps="DevTools → iPad",
         expected="Layout 2 cột cho product list, font đọc tốt", type="Compatibility Test"),
]


# ==================== SHOP / SEARCH & FILTER ====================
SHOP_CASES = [
    base(screen="Shop", feature="Layout", component="Page",       object="Layout trang /shop",
         pre="Truy cập GET /shop",
         steps="Vào /shop",
         expected="Hiển thị: breadcrumb, sidebar filter, sort dropdown, list sản phẩm, phân trang", type="UI"),
    base(screen="Shop", feature="Search bar", component="Input",  object="Hiển thị ô tìm kiếm",
         steps="Quan sát ô search trên header",
         expected="Input + icon kính lúp, gửi GET /shop?keyword=...", type="UI"),
    base(screen="Shop", feature="Filter sidebar", component="Category list", object="Sidebar lọc danh mục",
         pre="DB có nhiều category",
         steps="Quan sát sidebar trái",
         expected="Hiển thị các category, có số sản phẩm bên cạnh, checkbox chọn được", type="UI"),
    base(screen="Shop", feature="Filter sidebar", component="Price range", object="Lọc theo khoảng giá",
         steps="Quan sát filter khoảng giá",
         expected="Hiển thị input min-max hoặc slider giá", type="UI"),
    base(screen="Shop", feature="Sort", component="Dropdown sort", object="Hiển thị tùy chọn sort",
         steps="Click dropdown sort",
         expected="Hiển thị: Mới nhất, Giá tăng dần, Giá giảm dần, Bán chạy", type="UI"),
    base(screen="Shop", feature="No result", component="Empty state", object="Tìm không có kết quả",
         steps="Search 'asdkjasdkjasd'",
         expected="Thông báo 'Không tìm thấy sản phẩm', gợi ý sản phẩm bán chạy", type="UI"),

    base(screen="Shop", feature="Search", component="Theo tên", object="Tìm tên sản phẩm",
         steps="1. Nhập 'hồng'\n2. Submit",
         expected="Hiển thị các sản phẩm có name chứa từ 'hồng'"),
    base(screen="Shop", feature="Search", component="Không phân biệt hoa thường", object="Case-insensitive",
         steps="Search 'HỒNG' và 'hồng'",
         expected="Trả về cùng kết quả"),
    base(screen="Shop", feature="Filter", component="Theo danh mục", object="Lọc theo category",
         pre="Có ít nhất 1 sản phẩm trong category",
         steps="Click checkbox 1 category trong sidebar",
         expected="URL có query ?category=... và list lọc đúng"),
    base(screen="Shop", feature="Filter", component="Theo khoảng giá", object="Lọc giá",
         steps="Nhập min=100000, max=500000",
         expected="Chỉ hiển thị sản phẩm có price trong khoảng"),
    base(screen="Shop", feature="Sort", component="Giá tăng dần", object="Sort price asc",
         steps="Chọn sort = 'Giá tăng dần'",
         expected="Sản phẩm sắp xếp theo price ASC"),
    base(screen="Shop", feature="Sort", component="Giá giảm dần", object="Sort price desc",
         steps="Chọn sort = 'Giá giảm dần'",
         expected="Sản phẩm sắp xếp theo price DESC"),
    base(screen="Shop", feature="Sort", component="Mới nhất", object="Sort newest",
         steps="Chọn sort = 'Mới nhất'",
         expected="Sắp xếp theo created_at DESC"),
    base(screen="Shop", feature="Pagination", component="Phân trang", object="Chuyển trang",
         pre="Tổng > 12 sản phẩm",
         steps="Click trang 2",
         expected="URL ?page=2, hiển thị batch tiếp theo, vẫn giữ filter hiện tại"),
    base(screen="Shop", feature="Combo", component="Filter + Sort + Search", object="Kết hợp",
         steps="1. Search 'hoa'\n2. Filter category=Bó hoa\n3. Sort giá tăng",
         expected="Kết quả thỏa mãn cả 3 điều kiện, URL có đủ query parameters"),
    base(screen="Shop", feature="Reset", component="Xoá filter", object="Reset filter",
         pre="Đang có filter active",
         steps="Click 'Xoá bộ lọc'",
         expected="URL trở về /shop sạch, hiển thị toàn bộ sản phẩm"),
    base(screen="Shop", feature="Product click", component="Card click", object="Mở chi tiết",
         steps="Click vào 1 sản phẩm",
         expected="Điều hướng /product/{slug}"),

    base(screen="Shop", feature="Search", component="Keyword rỗng", object="Submit search rỗng",
         steps="Submit search với ô trống",
         expected="Trả về toàn bộ /shop (không lọc gì)", type="Abnormal"),
    base(screen="Shop", feature="Search", component="Quá dài", object="Keyword > 255 ký tự",
         steps="Nhập 300 ký tự lặp",
         expected="Validate cắt còn 255 hoặc báo lỗi, không 500", type="Abnormal"),
    base(screen="Shop", feature="Filter", component="Price âm",  object="min < 0",
         steps="Nhập min=-100, max=500000",
         expected="Coi min=0, không sinh truy vấn lỗi, không 500", type="Abnormal"),
    base(screen="Shop", feature="Filter", component="Min > Max", object="Hoán đổi",
         steps="min=900000, max=100000",
         expected="Báo lỗi hoặc tự hoán đổi, không 500", type="Abnormal"),
    base(screen="Shop", feature="Pagination", component="Trang quá lớn", object="?page=9999",
         steps="Truy cập ?page=9999",
         expected="Hiển thị trang trống, không 500", type="Abnormal"),
    base(screen="Shop", feature="Category", component="Category ẩn", object="Slug không tồn tại",
         steps="/shop?category=khong-ton-tai",
         expected="Hiển thị thông báo không có sản phẩm, không 500", type="Abnormal"),

    base(screen="Shop", feature="DB sync", component="Tổng sản phẩm", object="Trùng DB",
         pre="DB có N sản phẩm is_active=1",
         steps="Đếm tổng số trên /shop khi không filter",
         expected="Tổng = N (đếm trên tất cả trang qua phân trang)", type="Data Integrity"),
    base(screen="Shop", feature="DB sync", component="Bộ lọc category", object="Số sản phẩm khớp DB",
         pre="DB: category X có k sản phẩm active",
         steps="Filter category=X",
         expected="Số sản phẩm hiển thị = k", type="Data Integrity"),
    base(screen="Shop", feature="DB sync", component="Sản phẩm inactive", object="Ẩn khỏi shop",
         pre="DB có sản phẩm is_active=0",
         steps="Mở /shop",
         expected="Sản phẩm is_active=0 KHÔNG xuất hiện", type="Data Integrity"),

    base(screen="Shop", feature="Security", component="SQLi", object="Inject vào keyword",
         steps="Search ' OR 1=1 --",
         expected="Eloquent ràng buộc bind, không lộ dữ liệu, kết quả như search literal",
         type="Access Control & Security"),
    base(screen="Shop", feature="Security", component="XSS reflective", object="Inject <script> vào search",
         steps="Search <script>alert(1)</script>",
         expected="Blade escape, in ra dưới dạng literal", type="Access Control & Security"),

    base(screen="Shop", feature="Performance", component="Index DB", object="Query response time",
         pre="DB có 500+ sản phẩm",
         steps="Search và filter, đo thời gian server",
         expected="Mỗi response < 800ms (đã đánh index trên category_id, is_active)",
         type="Performance Test"),
    base(screen="Shop", feature="Performance", component="N+1 query", object="Kiểm tra N+1",
         steps="Bật Laravel Debugbar, mở /shop?page=1",
         expected="Số query < 10 (đã dùng with('category'))", type="Performance Test",
         note="Trước fix lượt query > 50; đã eager-load category & images"),

    base(screen="Shop", feature="Browser", component="Chrome", object="Test trên Chrome",
         steps="Mở /shop", expected="Hoạt động đúng", type="Compatibility Test"),
    base(screen="Shop", feature="Browser", component="Edge",   object="Test trên Edge",
         steps="Mở /shop", expected="Hoạt động đúng", type="Compatibility Test"),
    base(screen="Shop", feature="Responsive", component="Mobile", object="Mobile 375px",
         steps="DevTools iPhone SE", expected="Sidebar filter biến thành drawer/modal", type="Compatibility Test"),
    base(screen="Shop", feature="Responsive", component="Tablet", object="Tablet 768px",
         steps="DevTools iPad", expected="Grid 2 cột, sidebar vẫn hiển thị", type="Compatibility Test"),
]


# ==================== PRODUCT DETAIL ====================
PRODUCT_CASES = [
    base(screen="Product Detail", feature="Layout", component="Page", object="Toàn bộ trang chi tiết",
         pre="Sản phẩm tồn tại, slug hợp lệ",
         steps="Truy cập /product/{slug}",
         expected="Layout: ảnh lớn, ảnh nhỏ, tên, giá, mô tả, size/color, nút Thêm vào giỏ, mua ngay, đánh giá", type="UI"),
    base(screen="Product Detail", feature="Gallery", component="Ảnh chính", object="Hiển thị ảnh lớn",
         steps="Quan sát ảnh chính",
         expected="Ảnh đúng tỉ lệ, không vỡ", type="UI"),
    base(screen="Product Detail", feature="Pricing", component="Giá", object="Hiển thị giá",
         steps="Quan sát giá",
         expected="Format VND đúng, in đậm, màu nhấn", type="UI"),
    base(screen="Product Detail", feature="Variant", component="Size", object="Chọn size",
         pre="Sản phẩm có nhiều size",
         steps="Click chọn size M",
         expected="Size được highlight, sẵn sàng thêm vào giỏ", type="UI"),
    base(screen="Product Detail", feature="Variant", component="Color", object="Chọn màu",
         pre="Sản phẩm có nhiều màu",
         steps="Click chọn 1 màu",
         expected="Màu được highlight", type="UI"),
    base(screen="Product Detail", feature="Quantity", component="Stepper", object="Tăng/giảm số lượng",
         steps="Click + và -",
         expected="Số lượng thay đổi, không cho < 1", type="UI"),
    base(screen="Product Detail", feature="Description", component="Tab mô tả", object="Hiển thị tab",
         steps="Click tab 'Mô tả'",
         expected="Hiển thị mô tả chi tiết sản phẩm (rich text)", type="UI"),
    base(screen="Product Detail", feature="Reviews",   component="Tab đánh giá", object="Hiển thị tab review",
         steps="Click tab 'Đánh giá'",
         expected="Hiển thị danh sách review (rating, tên user, nội dung)", type="UI"),
    base(screen="Product Detail", feature="Related products", component="Section liên quan", object="Sản phẩm cùng danh mục",
         steps="Cuộn xuống cuối trang",
         expected="Hiển thị 4-8 sản phẩm cùng category", type="UI"),

    base(screen="Product Detail", feature="Add to cart", component="Nút", object="Thêm vào giỏ thành công",
         pre="Sản phẩm còn hàng",
         steps="1. Chọn size, color, quantity=2\n2. Click 'Thêm vào giỏ'",
         expected="Toast 'Đã thêm', badge cart +2, cart lưu đúng size/color/qty"),
    base(screen="Product Detail", feature="Buy now",     component="Nút", object="Mua ngay",
         pre="User đã đăng nhập, sản phẩm còn hàng",
         steps="Click 'Mua ngay'",
         expected="Chuyển đến /checkout với đúng sản phẩm + variant"),
    base(screen="Product Detail", feature="Add to cart", component="Chưa login", object="Guest add cart",
         pre="User chưa đăng nhập",
         steps="Click 'Thêm vào giỏ'",
         expected="Cart lưu vào session, vẫn cộng badge"),
    base(screen="Product Detail", feature="Wishlist",    component="Yêu thích đã login", object="Thêm yêu thích",
         pre="User đã đăng nhập",
         steps="Click icon trái tim",
         expected="Icon filled, toast 'Đã thêm vào yêu thích', wishlists table có record"),
    base(screen="Product Detail", feature="Reviews", component="Submit đánh giá", object="Gửi review thành công",
         pre="User đã đăng nhập + đã mua sản phẩm",
         steps="1. Click 'Viết đánh giá'\n2. Chọn 5 sao + nội dung\n3. Submit",
         expected="Review lưu DB với is_visible=1, hiển thị trên trang"),
    base(screen="Product Detail", feature="Reviews", component="Trung bình rating", object="Tính lại trung bình",
         pre="Sản phẩm có 3 review 5,4,3",
         steps="Quan sát rating tổng",
         expected="Hiển thị 4.0 sao + 3 đánh giá"),

    base(screen="Product Detail", feature="Stock", component="Hết hàng", object="Stock = 0",
         pre="Sản phẩm stock=0",
         steps="Quan sát nút 'Thêm vào giỏ'",
         expected="Nút disable + ghi rõ 'Hết hàng'", type="Abnormal"),
    base(screen="Product Detail", feature="Stock", component="Đặt quá tồn", object="qty > stock",
         pre="Sản phẩm có stock=3",
         steps="Tăng qty lên 10, thêm vào giỏ",
         expected="Báo lỗi 'Số lượng vượt tồn kho', không thêm vào giỏ", type="Abnormal",
         note="Trước đó cho phép vượt - đã fix tại CartController@add"),
    base(screen="Product Detail", feature="Slug",  component="Slug sai", object="Slug không tồn tại",
         steps="/product/khong-ton-tai",
         expected="HTTP 404 page", type="Abnormal"),
    base(screen="Product Detail", feature="Sản phẩm ẩn", component="is_active=0", object="Truy cập sản phẩm ẩn",
         pre="Admin set is_active=0",
         steps="/product/{slug}",
         expected="HTTP 404 (route binding filter)", type="Abnormal"),
    base(screen="Product Detail", feature="Reviews", component="Chưa login", object="Submit review khi guest",
         pre="User chưa đăng nhập",
         steps="POST /reviews",
         expected="Redirect /login", type="Abnormal"),
    base(screen="Product Detail", feature="Reviews", component="Chưa mua hàng", object="User chưa mua",
         pre="User đã login, chưa mua sản phẩm",
         steps="Cố vào /reviews/create/{slug}",
         expected="Báo lỗi 'Bạn cần mua sản phẩm trước khi đánh giá'", type="Abnormal"),
    base(screen="Product Detail", feature="Reviews", component="Rating ngoài 1-5", object="Submit rating=10",
         steps="POST /reviews với rating=10",
         expected="Validate fail, trả về form với lỗi", type="Abnormal"),
    base(screen="Product Detail", feature="Variant", component="Bỏ qua size", object="Chưa chọn size bắt buộc",
         pre="Sản phẩm yêu cầu chọn size",
         steps="Click thêm vào giỏ mà chưa chọn size",
         expected="Báo 'Vui lòng chọn size'", type="Abnormal"),

    base(screen="Product Detail", feature="DB sync", component="Giá", object="Giá hiển thị == DB",
         steps="So sánh giá trên trang vs price trong DB",
         expected="Hoàn toàn trùng khớp", type="Data Integrity"),
    base(screen="Product Detail", feature="DB sync", component="Stock", object="Stock = tồn kho DB",
         steps="So sánh tồn trên trang vs DB.products.stock",
         expected="Trùng khớp", type="Data Integrity"),
    base(screen="Product Detail", feature="DB sync", component="Review count", object="Đếm review",
         steps="Đếm review hiển thị vs DB",
         expected="Trùng khớp (chỉ tính is_visible=1)", type="Data Integrity"),

    base(screen="Product Detail", feature="Security", component="XSS trong review", object="Nội dung review XSS",
         steps="Submit review nội dung '<script>alert(1)</script>'",
         expected="Render thành literal, không thực thi JS", type="Access Control & Security"),
    base(screen="Product Detail", feature="Security", component="CSRF", object="Submit add-to-cart không có token",
         steps="POST /cart/add không có _token",
         expected="HTTP 419", type="Access Control & Security"),

    base(screen="Product Detail", feature="Performance", component="LCP", object="LCP đo bằng Lighthouse",
         steps="Lighthouse /product/{slug}",
         expected="LCP < 3.0s", type="Performance Test"),
    base(screen="Product Detail", feature="Performance", component="Spam buy-now", object="Click mua ngay liên tục",
         steps="Click 'Mua ngay' 5 lần liên tục",
         expected="Chỉ tạo 1 phiên checkout (debounce hoặc cùng order_id chờ)", type="Performance Test"),

    base(screen="Product Detail", feature="Browser", component="Chrome", object="OK trên Chrome",
         steps="Mở /product/{slug}", expected="OK", type="Compatibility Test"),
    base(screen="Product Detail", feature="Browser", component="Firefox", object="OK trên Firefox",
         steps="Mở /product/{slug}", expected="OK", type="Compatibility Test"),
    base(screen="Product Detail", feature="Responsive", component="Mobile", object="Mobile 375px",
         steps="DevTools iPhone SE",
         expected="Ảnh trên, info dưới, không scroll ngang", type="Compatibility Test"),
    base(screen="Product Detail", feature="Responsive", component="Tablet", object="Tablet 768px",
         steps="DevTools iPad",
         expected="Layout 2 cột, tab review hiển thị tốt", type="Compatibility Test"),
]


# ==================== CART ====================
CART_CASES = [
    base(screen="Cart", feature="Layout", component="Page", object="Trang /cart",
         pre="Có ít nhất 1 sản phẩm trong giỏ",
         steps="Truy cập /cart",
         expected="Bảng giỏ hàng: ảnh, tên, variant, đơn giá, qty, thành tiền, tổng, ô voucher, nút thanh toán", type="UI"),
    base(screen="Cart", feature="Empty",  component="Trạng thái rỗng", object="Giỏ trống",
         pre="Cart rỗng",
         steps="Truy cập /cart",
         expected="Thông báo 'Giỏ hàng trống' + nút 'Tiếp tục mua sắm' → /shop", type="UI"),
    base(screen="Cart", feature="Mini cart", component="Dropdown",   object="Mini cart trên header",
         pre="Có ≥ 1 sản phẩm",
         steps="Hover/Click icon giỏ",
         expected="Dropdown hiển thị 3 sản phẩm gần nhất + Subtotal + nút Xem giỏ", type="UI"),

    base(screen="Cart", feature="Add", component="Thêm sản phẩm mới", object="Add lần đầu",
         pre="Sản phẩm chưa có trong cart",
         steps="POST /cart/add với product_id và qty=1",
         expected="Tạo record carts, badge cart +1"),
    base(screen="Cart", feature="Add", component="Sản phẩm đã có", object="Gộp số lượng",
         pre="Sản phẩm đã có trong cart, qty=1",
         steps="POST /cart/add cùng product+variant, qty=1",
         expected="Update qty thành 2 (không tạo record mới)"),
    base(screen="Cart", feature="Update", component="Thay đổi số lượng", object="Update qty",
         pre="Có record cart",
         steps="POST /cart/{id}/update với qty=3",
         expected="DB cập nhật qty=3, tổng tiền tính lại"),
    base(screen="Cart", feature="Remove", component="Xoá khỏi cart", object="Remove item",
         pre="Có record cart",
         steps="DELETE /cart/{id}/remove",
         expected="Record bị xoá, badge cart -1"),
    base(screen="Cart", feature="Clear",  component="Xoá toàn bộ",   object="Clear cart",
         pre="Có nhiều sản phẩm",
         steps="POST /cart/clear",
         expected="Toàn bộ record bị xoá"),
    base(screen="Cart", feature="Voucher", component="Apply voucher", object="Áp voucher hợp lệ",
         pre="Voucher còn hiệu lực, đủ điều kiện",
         steps="Nhập code và Apply",
         expected="Subtotal giảm theo voucher, hiển thị 'Đã áp dụng'"),
    base(screen="Cart", feature="Voucher", component="Tính lại", object="Đổi số lượng sau khi áp voucher",
         pre="Đã áp voucher",
         steps="Update qty",
         expected="Discount tính lại theo tổng mới, không giảm sai"),
    base(screen="Cart", feature="Subtotal", component="Tổng tiền", object="Đúng = sum(price*qty)",
         pre="Có nhiều sản phẩm",
         steps="Quan sát Subtotal",
         expected="= SUM(price * qty) của tất cả item"),

    base(screen="Cart", feature="Add", component="Stock 0", object="Thêm sản phẩm hết hàng",
         pre="Sản phẩm stock=0",
         steps="POST /cart/add",
         expected="Báo lỗi 'Hết hàng', không thêm", type="Abnormal"),
    base(screen="Cart", feature="Add", component="qty > stock", object="Vượt tồn kho",
         pre="Stock=3, gửi qty=10",
         steps="POST /cart/add qty=10",
         expected="Báo lỗi 'Số lượng vượt tồn kho', cart không tăng", type="Abnormal",
         note="Trước đó cho thêm sai - đã fix CartController@add"),
    base(screen="Cart", feature="Update", component="qty=0", object="Đặt qty=0",
         pre="Có record cart",
         steps="POST /cart/{id}/update qty=0",
         expected="Auto xoá record (qty=0 = remove)", type="Abnormal"),
    base(screen="Cart", feature="Update", component="qty âm", object="qty=-1",
         steps="POST /cart/{id}/update qty=-1",
         expected="Validate fail, không cập nhật", type="Abnormal"),
    base(screen="Cart", feature="Voucher", component="Code không tồn tại", object="Voucher sai",
         steps="Apply 'WRONGCODE'",
         expected="Báo 'Mã không hợp lệ'", type="Abnormal"),
    base(screen="Cart", feature="Voucher", component="Hết hạn", object="Voucher expired",
         pre="Voucher quá end_date",
         steps="Apply code expired",
         expected="Báo 'Mã đã hết hạn'", type="Abnormal"),
    base(screen="Cart", feature="Voucher", component="Đơn không đủ min", object="Subtotal < min_order_amount",
         steps="Apply voucher yêu cầu tối thiểu 500k khi cart 100k",
         expected="Báo 'Đơn tối thiểu chưa đạt'", type="Abnormal"),
    base(screen="Cart", feature="Voucher", component="Đã sử dụng", object="User đã dùng",
         pre="User đã dùng voucher 1 lần với usage_limit_per_user=1",
         steps="Apply lại",
         expected="Báo 'Mã đã được sử dụng'", type="Abnormal"),

    base(screen="Cart", feature="Persist", component="Guest cart", object="Giỏ ẩn theo session",
         pre="Guest có item trong cart",
         steps="Đóng browser → mở lại trong 1h",
         expected="Cart vẫn còn (session cookie chưa hết)", type="Data Integrity"),
    base(screen="Cart", feature="Persist", component="Merge khi login", object="Merge guest cart vào user cart",
         pre="Guest có 2 item, user đã có 1 item",
         steps="Login",
         expected="User cart có cả 3 item (gộp nếu trùng product+variant)", type="Data Integrity"),
    base(screen="Cart", feature="Persist", component="DB consistent", object="Giá lưu trong cart",
         pre="Giá sản phẩm đã đổi từ 100k → 120k sau khi user thêm vào giỏ",
         steps="Mở /cart",
         expected="Cart hiển thị giá hiện hành (snapshot có thể là 120k - tùy spec)", type="Data Integrity"),

    base(screen="Cart", feature="Security", component="Truy cập cart user khác", object="Xem giỏ user khác",
         pre="A có cart record id=5",
         steps="B request POST /cart/5/update",
         expected="HTTP 403 hoặc cart không thuộc B → ignore", type="Access Control & Security",
         note="Bổ sung policy kiểm tra user_id match"),
    base(screen="Cart", feature="Security", component="CSRF", object="Không token",
         steps="POST /cart/add không có _token",
         expected="HTTP 419", type="Access Control & Security"),

    base(screen="Cart", feature="Performance", component="Trên 50 items", object="Cart lớn",
         pre="Cart có 50 items",
         steps="Mở /cart",
         expected="< 1s render", type="Performance Test"),
    base(screen="Cart", feature="Performance", component="Spam add", object="Click add 10 lần",
         steps="Click 'Thêm vào giỏ' 10 lần",
         expected="Debounce, không tạo 10 request thành công", type="Performance Test"),

    base(screen="Cart", feature="Browser", component="Chrome",  object="OK Chrome",
         steps="Mở /cart", expected="OK", type="Compatibility Test"),
    base(screen="Cart", feature="Browser", component="Firefox", object="OK Firefox",
         steps="Mở /cart", expected="OK", type="Compatibility Test"),
    base(screen="Cart", feature="Responsive", component="Mobile", object="Mobile 375px",
         steps="DevTools",
         expected="Card item dọc, vẫn dễ thao tác", type="Compatibility Test"),
    base(screen="Cart", feature="Responsive", component="Tablet", object="Tablet 768px",
         steps="DevTools",
         expected="Bảng cart vẫn hiển thị đúng", type="Compatibility Test"),
]


# ==================== CHECKOUT & PAYMENT ====================
CHECKOUT_CASES = [
    base(screen="Checkout", feature="Layout", component="Page", object="Layout /checkout",
         pre="User đã login, có ít nhất 1 item cart",
         steps="Truy cập /checkout",
         expected="Form: tên, sđt, địa chỉ, ghi chú; tóm tắt đơn; ô voucher; chọn ship; chọn thanh toán; nút đặt hàng", type="UI"),
    base(screen="Checkout", feature="Order summary", component="Tóm tắt", object="Hiển thị thông tin đơn",
         steps="Quan sát section bên phải",
         expected="Liệt kê sản phẩm + variant + qty + tổng tiền + phí ship + giảm giá", type="UI"),
    base(screen="Checkout", feature="Voucher", component="Available list", object="Danh sách voucher",
         steps="Click 'Chọn mã' / 'Available vouchers'",
         expected="GET /checkout/available-vouchers trả JSON list voucher đủ điều kiện", type="UI"),
    base(screen="Checkout", feature="Payment method", component="Chọn thanh toán", object="Hiển thị 2 phương thức",
         steps="Quan sát",
         expected="Có 2 lựa chọn: COD (Cash on Delivery) và Thẻ ngân hàng (Card)", type="UI"),
    base(screen="Checkout", feature="Shipping", component="Estimate", object="Tính phí ship",
         pre="Đã nhập địa chỉ",
         steps="Submit POST /shipping/estimate",
         expected="Trả về phí ship dựa trên tỉnh/thành"),

    base(screen="Checkout", feature="Place order", component="COD", object="Đặt hàng thành công COD",
         pre="Form hợp lệ",
         steps="1. Điền form\n2. Chọn COD\n3. Click 'Đặt hàng'",
         expected="Tạo order status=pending, redirect /checkout/success, gửi mail (nếu cấu hình SMTP), cart bị xoá"),
    base(screen="Checkout", feature="Place order", component="Card", object="Đặt hàng thẻ",
         pre="Form hợp lệ",
         steps="1. Điền form\n2. Chọn 'Thẻ'\n3. Đặt hàng",
         expected="Tạo order status=awaiting_payment, redirect /checkout/card/{order}"),
    base(screen="Checkout", feature="Card payment", component="Confirm card", object="Xác nhận thanh toán thẻ",
         pre="Đang ở /checkout/card/{order}",
         steps="Nhập thông tin thẻ mock, click 'Confirm'",
         expected="Order status=paid, redirect /checkout/success"),
    base(screen="Checkout", feature="Voucher", component="Apply preview", object="Apply voucher tại checkout",
         pre="Voucher hợp lệ",
         steps="POST /checkout/apply-voucher",
         expected="Trả JSON: discount, new total, ok=true"),
    base(screen="Checkout", feature="Stock deduction", component="Trừ kho", object="Stock giảm khi place order",
         pre="Sản phẩm stock=10, order qty=3",
         steps="Place order thành công",
         expected="Stock còn 7, orders.stock_deducted=1"),
    base(screen="Checkout", feature="Buy now", component="Init buy now", object="Mua ngay 1 sản phẩm",
         pre="User đã login",
         steps="POST /checkout/buy-now với product_id, qty, variant",
         expected="Bypass cart, vào /checkout với chỉ sản phẩm đó"),
    base(screen="Checkout", feature="Success page", component="Confirmation", object="/checkout/success",
         pre="Vừa place order",
         steps="Quan sát trang success",
         expected="Hiển thị order_id, tổng tiền, hướng dẫn theo phương thức thanh toán"),

    base(screen="Checkout", feature="Validation", component="Thiếu tên", object="Bỏ trống tên",
         steps="Submit không có name",
         expected="Báo 'Vui lòng nhập tên'", type="Abnormal"),
    base(screen="Checkout", feature="Validation", component="SĐT sai định dạng", object="Phone invalid",
         steps="Nhập phone='abc'",
         expected="Báo 'Số điện thoại không hợp lệ'", type="Abnormal"),
    base(screen="Checkout", feature="Validation", component="Address rỗng", object="Bỏ địa chỉ",
         steps="Submit không có address",
         expected="Báo 'Vui lòng nhập địa chỉ'", type="Abnormal"),
    base(screen="Checkout", feature="Validation", component="Cart rỗng", object="Vào checkout khi cart rỗng",
         pre="Cart rỗng",
         steps="Truy cập /checkout",
         expected="Redirect /cart hoặc /shop với flash 'Giỏ hàng trống'", type="Abnormal"),
    base(screen="Checkout", feature="Stock", component="Out of stock khi đặt", object="Stock vừa hết",
         pre="Admin set stock=0 khi user đang ở /checkout",
         steps="Click 'Đặt hàng'",
         expected="Báo lỗi 'Sản phẩm vừa hết hàng', không tạo order", type="Abnormal"),
    base(screen="Checkout", feature="Voucher", component="Voucher invalid trong checkout", object="Voucher hết hạn",
         steps="Apply voucher hết hạn",
         expected="Báo 'Mã không hợp lệ', subtotal không thay đổi", type="Abnormal"),
    base(screen="Checkout", feature="Network", component="Mất kết nối", object="Submit lúc mất mạng",
         steps="Tắt mạng → Submit",
         expected="Thông báo lỗi rõ ràng, không double tạo order", type="Abnormal"),

    base(screen="Checkout", feature="DB", component="Order tổng tiền", object="Tổng = sum item + ship - discount",
         pre="Vừa place order thành công",
         steps="So sánh orders.total_amount vs sum(order_items.subtotal) + ship - discount",
         expected="Trùng khớp", type="Data Integrity"),
    base(screen="Checkout", feature="DB", component="Order items snapshot", object="Ghi snapshot giá",
         pre="Vừa place order",
         steps="So order_items.price vs products.price tại thời điểm đặt",
         expected="Bằng nhau (snapshot)", type="Data Integrity"),
    base(screen="Checkout", feature="DB", component="Voucher usage", object="VoucherUserUsage tăng",
         pre="Đặt hàng với voucher",
         steps="Đếm record voucher_user_usage",
         expected="+1 record với đúng user_id, voucher_id", type="Data Integrity"),
    base(screen="Checkout", feature="DB", component="Stock deducted flag", object="Trừ kho 1 lần",
         steps="Place order, sau đó hủy, lại place lại",
         expected="Mỗi order chỉ trừ kho 1 lần, hủy thì hoàn lại", type="Data Integrity"),

    base(screen="Checkout", feature="Auth", component="Yêu cầu login", object="Guest checkout",
         pre="User chưa login",
         steps="GET /checkout",
         expected="Redirect /login với ?redirect_to=/checkout", type="Access Control & Security"),
    base(screen="Checkout", feature="CSRF", component="POST không token", object="CSRF check",
         steps="POST /checkout không có _token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="Checkout", feature="Authz", component="Xem order user khác", object="Truy cập /checkout/success?order=X của user khác",
         pre="A có order id=10",
         steps="B request /checkout/success?order=10",
         expected="HTTP 403", type="Access Control & Security",
         note="Đã thêm Gate kiểm tra ownership"),
    base(screen="Checkout", feature="Card", component="SSL", object="Trang thẻ ép HTTPS",
         steps="Truy cập http://.../checkout/card/X",
         expected="Redirect HTTPS, hoặc cảnh báo bảo mật", type="Access Control & Security"),

    base(screen="Checkout", feature="Performance", component="Double submit", object="Submit nhanh 2 lần",
         steps="Click 'Đặt hàng' liên tục 2 lần",
         expected="Chỉ tạo 1 order (idempotency token / disable nút)", type="Performance Test",
         note="Trước đó tạo 2 order trùng; đã fix bằng disable nút sau submit"),
    base(screen="Checkout", feature="Performance", component="Race trừ kho", object="2 user cùng đặt 1 sản phẩm còn 1",
         pre="Stock=1",
         steps="2 trình duyệt cùng click 'Đặt hàng'",
         expected="Chỉ 1 thành công (DB transaction + lock); user còn lại nhận 'Hết hàng'", type="Performance Test"),

    base(screen="Checkout", feature="Browser", component="Chrome", object="OK Chrome",
         steps="Place test order", expected="OK", type="Compatibility Test"),
    base(screen="Checkout", feature="Browser", component="Firefox", object="OK Firefox",
         steps="Place test order", expected="OK", type="Compatibility Test"),
    base(screen="Checkout", feature="Responsive", component="Mobile", object="Mobile 375px",
         steps="DevTools",
         expected="Form chiếm full width, tóm tắt nằm phía trên hoặc collapsable", type="Compatibility Test"),
    base(screen="Checkout", feature="Responsive", component="Tablet", object="Tablet 768px",
         steps="DevTools",
         expected="Layout 2 cột giữ vững", type="Compatibility Test"),
]


# ==================== LOGIN ====================
LOGIN_CASES = [
    base(screen="Login", feature="Layout", component="Page", object="Trang /login",
         pre="User chưa đăng nhập",
         steps="Truy cập /login",
         expected="Form: email, password, remember me, nút Đăng nhập, link 'Đăng ký', 'Quên mật khẩu'", type="UI"),
    base(screen="Login", feature="Email field", component="Input", object="Hiển thị email",
         steps="Quan sát ô email",
         expected="Type=email, placeholder='Email'", type="UI"),
    base(screen="Login", feature="Password field", component="Input", object="Hiển thị password",
         steps="Quan sát ô password",
         expected="Type=password, ký tự bị che", type="UI"),
    base(screen="Login", feature="Remember me", component="Checkbox", object="Hiển thị remember me",
         steps="Quan sát checkbox",
         expected="Có checkbox 'Ghi nhớ đăng nhập'", type="UI"),

    base(screen="Login", feature="Login OK", component="Customer login", object="Login user thường",
         pre="DB có user minhanh@gmail.com / password",
         steps="1. Nhập email + password\n2. Submit",
         expected="Đăng nhập thành công, redirect / hoặc trang trước, session lưu user_id"),
    base(screen="Login", feature="Remember me", component="Persist", object="Cookie remember",
         pre="Tick remember me",
         steps="1. Login\n2. Đóng browser 1h\n3. Mở lại",
         expected="Vẫn đăng nhập (remember_token được set)"),
    base(screen="Login", feature="Admin login", component="Admin auth", object="Login admin",
         pre="DB có admin admin@huonghoaxinh.com",
         steps="POST /admin/login với credential admin",
         expected="Redirect /admin/dashboard"),
    base(screen="Login", feature="Forgot password", component="Link", object="Click 'Quên mật khẩu'",
         steps="Click link",
         expected="Điều hướng /forgot-password"),

    base(screen="Login", feature="Validation", component="Email rỗng", object="Bỏ email",
         steps="Submit chỉ password",
         expected="Báo 'Vui lòng nhập email'", type="Abnormal"),
    base(screen="Login", feature="Validation", component="Password rỗng", object="Bỏ password",
         steps="Submit chỉ email",
         expected="Báo 'Vui lòng nhập mật khẩu'", type="Abnormal"),
    base(screen="Login", feature="Validation", component="Email sai format", object="Format invalid",
         steps="Nhập email='abc'",
         expected="Báo 'Email không hợp lệ'", type="Abnormal"),
    base(screen="Login", feature="Auth fail", component="Sai password", object="Sai mật khẩu",
         steps="Đúng email, sai password",
         expected="Báo 'Thông tin đăng nhập không đúng', không reveal email vs password", type="Abnormal"),
    base(screen="Login", feature="Auth fail", component="Email không tồn tại", object="Không tồn tại",
         steps="Nhập email không có trong DB",
         expected="Báo 'Thông tin đăng nhập không đúng' (chung chung)", type="Abnormal"),
    base(screen="Login", feature="Locked",   component="Tài khoản bị khóa", object="is_locked=1",
         pre="DB user có is_locked=1",
         steps="Login với credential đúng",
         expected="Báo 'Tài khoản đã bị khóa, liên hệ admin'", type="Abnormal",
         note="Đã bổ sung guard kiểm tra is_locked trong login()"),

    base(screen="Login", feature="DB", component="Session", object="Lưu session đúng user_id",
         steps="Login user A → check session",
         expected="session.user_id = A.id", type="Data Integrity"),
    base(screen="Login", feature="DB", component="Last login", object="Cập nhật last_login_at (nếu có)",
         steps="Login",
         expected="users.last_login_at = now()", type="Data Integrity"),

    base(screen="Login", feature="Security", component="SQLi email", object="' OR 1=1 --",
         steps="Email=\"' OR 1=1 --\"",
         expected="Validate hoặc Eloquent escape, không login as another user", type="Access Control & Security"),
    base(screen="Login", feature="Security", component="Brute force", object="5 lần sai liên tiếp",
         steps="Login sai 5 lần",
         expected="Throttle 1 phút (Laravel throttle middleware)", type="Access Control & Security",
         note="Mặc định throttle 5 attempts; verify hoạt động"),
    base(screen="Login", feature="Security", component="CSRF", object="POST không token",
         steps="POST /login không token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="Login", feature="Security", component="HTTPS", object="Login qua HTTPS",
         steps="Quan sát URL",
         expected="Production phải HTTPS", type="Access Control & Security"),
    base(screen="Login", feature="Security", component="Session fixation", object="Regen session sau login",
         steps="Login và so session id trước/sau",
         expected="Session ID thay đổi sau khi login (regenerate)", type="Access Control & Security"),

    base(screen="Login", feature="Performance", component="Time", object="Response < 500ms",
         steps="Đo thời gian POST /login",
         expected="< 500ms (server)", type="Performance Test"),

    base(screen="Login", feature="Browser",   component="Chrome",  object="OK", steps="Mở /login", expected="OK", type="Compatibility Test"),
    base(screen="Login", feature="Browser",   component="Firefox", object="OK", steps="Mở /login", expected="OK", type="Compatibility Test"),
    base(screen="Login", feature="Responsive",component="Mobile",  object="375px", steps="DevTools", expected="Form căn giữa, full width", type="Compatibility Test"),
    base(screen="Login", feature="Responsive",component="Tablet",  object="768px", steps="DevTools", expected="OK", type="Compatibility Test"),
]


# ==================== REGISTER ====================
REGISTER_CASES = [
    base(screen="Register", feature="Layout", component="Page", object="Trang /register",
         pre="User chưa đăng nhập",
         steps="Truy cập /register",
         expected="Form: name, email, password, confirm password, nút Đăng ký, link 'Đã có tài khoản'", type="UI"),
    base(screen="Register", feature="Password field", component="Toggle", object="Hiện/ẩn password",
         steps="Click icon mắt",
         expected="Toggle type=text/password", type="UI"),

    base(screen="Register", feature="Happy", component="Đăng ký thành công", object="Tạo user mới",
         pre="Email chưa có trong DB",
         steps="1. Nhập name, email mới, password\n2. Submit",
         expected="Tạo user, auto login, redirect /, gửi mail verify (nếu cấu hình)"),
    base(screen="Register", feature="Auto login", component="Sau register", object="Auto login",
         steps="Sau khi register",
         expected="session.user_id đã set"),

    base(screen="Register", feature="Validation", component="Name rỗng", object="Bỏ name",
         steps="Submit không có name",
         expected="Báo 'Vui lòng nhập tên'", type="Abnormal"),
    base(screen="Register", feature="Validation", component="Email rỗng", object="Bỏ email",
         steps="Submit không có email",
         expected="Báo 'Vui lòng nhập email'", type="Abnormal"),
    base(screen="Register", feature="Validation", component="Email sai format", object="abc.com",
         steps="Email='abc.com'",
         expected="Báo 'Email không hợp lệ'", type="Abnormal"),
    base(screen="Register", feature="Validation", component="Email trùng", object="Email đã tồn tại",
         pre="DB có email A",
         steps="Submit register với email A",
         expected="Báo 'Email đã tồn tại'", type="Abnormal"),
    base(screen="Register", feature="Validation", component="Password ngắn", object="< 8 ký tự",
         steps="Password='123'",
         expected="Báo 'Mật khẩu tối thiểu 8 ký tự'", type="Abnormal"),
    base(screen="Register", feature="Validation", component="Password không khớp confirm", object="Mismatch",
         steps="Password='12345678', confirm='87654321'",
         expected="Báo 'Mật khẩu xác nhận không khớp'", type="Abnormal"),

    base(screen="Register", feature="DB", component="Hash password", object="Lưu password đã hash",
         pre="Vừa register",
         steps="Mở DB",
         expected="users.password là bcrypt hash, không lưu plain", type="Data Integrity"),
    base(screen="Register", feature="DB", component="Role mặc định", object="role='user'",
         pre="Vừa register",
         steps="Check users.role",
         expected="role='user', is_locked=0", type="Data Integrity"),
    base(screen="Register", feature="DB", component="Email lowercase", object="Email lưu thường",
         steps="Register với Email='Test@Mail.com'",
         expected="users.email = 'test@mail.com'", type="Data Integrity",
         note="Bổ sung lower() trước save"),

    base(screen="Register", feature="Security", component="XSS name", object="Tên chứa <script>",
         steps="Name='<script>alert(1)</script>'",
         expected="Lưu literal, render escape, không thực thi", type="Access Control & Security"),
    base(screen="Register", feature="Security", component="CSRF", object="Không token",
         steps="POST /register không _token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="Register", feature="Security", component="Mass assignment", object="Inject role=admin",
         steps="POST /register với extra field role=admin",
         expected="role vẫn = 'user' (do $fillable không bao gồm role)", type="Access Control & Security",
         note="Đã kiểm tra $fillable - role không nằm trong mass-assignable"),
    base(screen="Register", feature="Security", component="Rate limit", object="Spam register",
         steps="POST /register 20 lần/phút",
         expected="Throttle hoạt động, trả 429", type="Access Control & Security"),

    base(screen="Register", feature="Performance", component="Time", object="Response < 800ms",
         steps="Đo thời gian register", expected="< 800ms", type="Performance Test"),

    base(screen="Register", feature="Browser",   component="Chrome",  object="OK", steps="Mở /register", expected="OK", type="Compatibility Test"),
    base(screen="Register", feature="Browser",   component="Firefox", object="OK", steps="Mở /register", expected="OK", type="Compatibility Test"),
    base(screen="Register", feature="Responsive",component="Mobile",  object="375px", steps="DevTools", expected="Form full width", type="Compatibility Test"),
    base(screen="Register", feature="Responsive",component="Tablet",  object="768px", steps="DevTools", expected="OK", type="Compatibility Test"),
]


# ==================== MY ACCOUNT ====================
ACCOUNT_CASES = [
    base(screen="My Account", feature="Profile", component="Page", object="Trang /profile",
         pre="User đã login",
         steps="Truy cập /profile",
         expected="Form chỉnh sửa: tên, email, sđt, địa chỉ, ngày sinh, nút Lưu", type="UI"),
    base(screen="My Account", feature="Avatar",  component="Upload",   object="Upload avatar",
         steps="Click 'Đổi avatar', chọn ảnh, submit",
         expected="Avatar được lưu, hiển thị mới", type="UI"),
    base(screen="My Account", feature="Sidebar", component="Menu account", object="Menu trái",
         steps="Quan sát sidebar",
         expected="Gồm: Hồ sơ, Đơn hàng, Yêu thích, Đánh giá, Đổi mật khẩu, Đăng xuất", type="UI"),
    base(screen="My Account", feature="Order history", component="Page", object="Trang /orders/history",
         steps="Click 'Đơn hàng của tôi'",
         expected="Bảng các đơn: mã, ngày, tổng, trạng thái, nút xem chi tiết / hủy / xác nhận nhận hàng", type="UI"),

    base(screen="My Account", feature="Profile update", component="Sửa hồ sơ", object="Update thông tin",
         steps="Sửa tên, lưu",
         expected="Lưu DB thành công, flash success"),
    base(screen="My Account", feature="Password change", component="Đổi mật khẩu", object="Đổi password",
         pre="Biết mật khẩu cũ",
         steps="Nhập password cũ + mới + xác nhận, submit",
         expected="Đổi thành công, mật khẩu mới hash trong DB"),
    base(screen="My Account", feature="Profile lock",  component="Tự khóa", object="Khóa tài khoản",
         steps="POST /profile/lock",
         expected="users.is_locked=1, logout, redirect /"),
    base(screen="My Account", feature="Profile destroy", component="Xoá tài khoản", object="Tự xoá",
         steps="DELETE /profile với password đúng",
         expected="User bị xoá hoặc soft delete, logout"),
    base(screen="My Account", feature="Order detail", component="Xem chi tiết", object="Click 1 đơn",
         steps="Click 'Xem' của 1 order",
         expected="Trang chi tiết: items, tổng, trạng thái, lịch sử trạng thái"),
    base(screen="My Account", feature="Order cancel", component="Hủy đơn", object="Hủy đơn pending",
         pre="Order status=pending",
         steps="Click 'Hủy đơn'",
         expected="Order status=cancelled, refund stock (orders.stock_deducted=0)"),
    base(screen="My Account", feature="Order confirm received", component="Đã nhận", object="Xác nhận nhận hàng",
         pre="Order status=shipping",
         steps="Click 'Đã nhận hàng'",
         expected="Order status=completed"),

    base(screen="My Account", feature="Validation", component="Email trùng khi update", object="Đổi sang email user khác",
         pre="Email B đã tồn tại",
         steps="Update email = B",
         expected="Báo 'Email đã được sử dụng'", type="Abnormal"),
    base(screen="My Account", feature="Validation", component="Đổi password sai cũ", object="Sai password cũ",
         steps="Nhập sai mật khẩu cũ",
         expected="Báo 'Mật khẩu cũ không đúng'", type="Abnormal"),
    base(screen="My Account", feature="Order cancel", component="Hủy đơn đã giao", object="Order status=shipping",
         pre="Order status=shipping",
         steps="Cố hủy",
         expected="Báo 'Không thể hủy đơn đang giao'", type="Abnormal"),
    base(screen="My Account", feature="Avatar", component="File quá lớn", object="> 2MB",
         steps="Upload ảnh 5MB",
         expected="Báo 'File quá lớn'", type="Abnormal"),
    base(screen="My Account", feature="Avatar", component="File sai loại", object="PDF",
         steps="Upload .pdf",
         expected="Báo 'Chỉ chấp nhận jpg/png'", type="Abnormal"),

    base(screen="My Account", feature="DB", component="Order list đúng user", object="Chỉ show order của user",
         steps="Vào /orders/history",
         expected="Chỉ liệt kê orders với user_id = current user", type="Data Integrity"),

    base(screen="My Account", feature="Auth", component="Yêu cầu login", object="Guest /profile",
         steps="GET /profile khi chưa login",
         expected="Redirect /login", type="Access Control & Security"),
    base(screen="My Account", feature="Authz", component="Xem order user khác", object="Order ID người khác",
         steps="GET /orders/{order} của user khác",
         expected="HTTP 403", type="Access Control & Security",
         note="Đã thêm policy OrderPolicy@view"),
    base(screen="My Account", feature="Security", component="CSRF profile update", object="Không token",
         steps="POST /profile không _token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="My Account", feature="Security", component="Avatar XSS", object="Upload file .php đổi đuôi",
         steps="Upload exploit.php.jpg",
         expected="Storage block, mime check thật sự", type="Access Control & Security"),

    base(screen="My Account", feature="Performance", component="Order list lớn", object="100 orders",
         steps="Mở /orders/history với user có 100 orders",
         expected="< 1s render (phân trang)", type="Performance Test"),

    base(screen="My Account", feature="Browser",   component="Chrome",  object="OK", steps="Mở /profile", expected="OK", type="Compatibility Test"),
    base(screen="My Account", feature="Browser",   component="Firefox", object="OK", steps="Mở /profile", expected="OK", type="Compatibility Test"),
    base(screen="My Account", feature="Responsive",component="Mobile",  object="375px", steps="DevTools", expected="Sidebar collapse, form full width", type="Compatibility Test"),
]


# ==================== WISHLIST & REVIEW ====================
WISHLIST_CASES = [
    base(screen="Wishlist", feature="Layout", component="Page", object="Trang /wishlist",
         pre="User đã login",
         steps="Truy cập /wishlist",
         expected="Danh sách sản phẩm yêu thích dạng grid", type="UI"),
    base(screen="Wishlist", feature="Empty",  component="Trạng thái rỗng", object="Wishlist trống",
         pre="User chưa thêm sản phẩm",
         steps="Truy cập /wishlist",
         expected="Hiển thị 'Bạn chưa có sản phẩm yêu thích nào' + nút 'Khám phá'", type="UI"),

    base(screen="Wishlist", feature="Add",    component="Toggle thêm", object="Thêm vào wishlist",
         pre="User đã login",
         steps="POST /wishlist/toggle với product_id",
         expected="Tạo record wishlists, button đổi sang trạng thái active"),
    base(screen="Wishlist", feature="Remove", component="Toggle bỏ", object="Bỏ khỏi wishlist",
         pre="Sản phẩm đã có trong wishlist",
         steps="POST /wishlist/toggle với product_id cũ",
         expected="Xoá record, button về trạng thái off"),
    base(screen="Wishlist", feature="Check",  component="API check", object="GET /wishlist/check/{id}",
         steps="Gọi API",
         expected="Trả JSON {in_wishlist: true/false}"),
    base(screen="Wishlist", feature="Add to cart from wishlist", component="Move", object="Chuyển sang giỏ",
         steps="Click 'Thêm vào giỏ' trong wishlist",
         expected="Add cart thành công, sản phẩm vẫn còn trong wishlist (nếu không yêu cầu xoá)"),

    base(screen="Wishlist", feature="Auth",   component="Yêu cầu login", object="Guest /wishlist",
         steps="GET /wishlist khi guest",
         expected="Redirect /login", type="Access Control & Security"),
    base(screen="Wishlist", feature="Authz",  component="Add cho user khác", object="Inject user_id",
         steps="POST với user_id=khác",
         expected="Bị ignore, lưu = auth user", type="Access Control & Security"),
    base(screen="Wishlist", feature="CSRF",   component="Không token", object="POST không _token",
         steps="POST /wishlist/add không token",
         expected="HTTP 419", type="Access Control & Security"),

    base(screen="Wishlist", feature="DB",     component="Unique constraint", object="Không trùng",
         pre="User đã có sản phẩm X",
         steps="POST add product X lần nữa",
         expected="Không tạo record trùng (unique (user_id, product_id))", type="Data Integrity"),

    base(screen="Review",   feature="List",   component="Liệt kê reviews",  object="Hiện trên /product/{slug}",
         pre="Sản phẩm có review is_visible=1",
         steps="Mở chi tiết sản phẩm",
         expected="Hiển thị các review, mới nhất trước", type="UI"),
    base(screen="Review",   feature="Submit", component="POST /reviews",     object="Submit thành công",
         pre="User đã mua sản phẩm",
         steps="Submit form review",
         expected="Tạo record, hiển thị ngay"),
    base(screen="Review",   feature="Validation", component="Rating ngoài 1-5", object="rating=6",
         steps="POST /reviews rating=6",
         expected="Validate fail", type="Abnormal"),
    base(screen="Review",   feature="Validation", component="Comment quá dài", object="> 2000 ký tự",
         steps="Submit comment 3000 ký tự",
         expected="Validate fail", type="Abnormal"),
    base(screen="Review",   feature="Authz",  component="Đánh giá khi chưa mua", object="User chưa mua sản phẩm",
         steps="GET /reviews/create/{slug}",
         expected="Báo 'Bạn cần mua sản phẩm trước'", type="Access Control & Security"),
    base(screen="Review",   feature="DB",     component="Average rating cập nhật", object="Tự cập nhật",
         pre="Vừa submit review",
         steps="Reload /product/{slug}",
         expected="Average rating tính lại bao gồm review mới", type="Data Integrity"),

    base(screen="Wishlist", feature="Performance", component="100 items", object="Trang lớn",
         steps="User có 100 wishlist",
         expected="< 1s render (paginate)", type="Performance Test"),
    base(screen="Wishlist", feature="Browser",     component="Chrome",  object="OK", steps="Mở /wishlist", expected="OK", type="Compatibility Test"),
    base(screen="Wishlist", feature="Responsive",  component="Mobile",  object="375px", steps="DevTools", expected="Grid 1 cột", type="Compatibility Test"),
]


# ==================== ADMIN DASHBOARD ====================
ADMIN_DASH_CASES = [
    base(screen="Admin Login", feature="Layout", component="Page", object="Trang /admin/login",
         pre="Chưa đăng nhập admin",
         steps="Truy cập /admin/login",
         expected="Form login admin riêng (route name admin.login)", type="UI"),
    base(screen="Admin Login", feature="Auth",   component="Login admin OK", object="Login OK",
         pre="DB có admin admin@huonghoaxinh.com / 12345678",
         steps="POST /admin/login",
         expected="Redirect /admin/dashboard"),
    base(screen="Admin Login", feature="Auth",   component="Sai password", object="Sai pass",
         steps="POST /admin/login với sai password",
         expected="Báo 'Thông tin đăng nhập không đúng'", type="Abnormal"),
    base(screen="Admin Login", feature="Auth",   component="User thường", object="Login bằng user thường",
         pre="email user thường",
         steps="POST /admin/login",
         expected="Không cho vào admin (chỉ admin guard mới được)", type="Access Control & Security"),

    base(screen="Admin Dashboard", feature="Layout", component="Page", object="Trang /admin/dashboard",
         pre="Đã login admin",
         steps="Truy cập /admin/dashboard",
         expected="Cards: tổng doanh thu, tổng đơn, tổng KH, tổng sản phẩm + biểu đồ doanh thu (Chart.js)", type="UI"),
    base(screen="Admin Dashboard", feature="Recent orders", component="Bảng", object="Bảng đơn gần đây",
         steps="Quan sát section gần đây",
         expected="5-10 đơn mới nhất, có badge NEW cho đơn chưa xem", type="UI"),
    base(screen="Admin Dashboard", feature="Chart doanh thu", component="Chart.js", object="Vẽ biểu đồ",
         steps="Quan sát chart",
         expected="Hiển thị doanh thu theo ngày/tháng, đầy đủ legend các status"),
    base(screen="Admin Dashboard", feature="Chart status", component="Legend", object="Hiển thị full status",
         steps="Quan sát legend",
         expected="Đầy đủ trạng thái orders, không bị cắt",
         note="Trước thiếu vài status; đã fix commit cb952ae"),
    base(screen="Admin Dashboard", feature="NEW badge", component="Order chưa xem", object="Limit 3 đơn",
         pre="DB có >3 order chưa xem viewed_at",
         steps="Quan sát danh sách",
         expected="Chỉ tối đa 3 đơn mới nhất có badge NEW",
         note="Fix commit c48c56f"),

    base(screen="Admin Dashboard", feature="Sidebar", component="Menu admin", object="Menu trái",
         steps="Quan sát sidebar",
         expected="Đủ mục: Dashboard, Sản phẩm, Danh mục, Đơn hàng, Khách hàng, Voucher, Đánh giá, Doanh thu, Tin nhắn, Cài đặt", type="UI"),
    base(screen="Admin Dashboard", feature="Doanh thu", component="Trang /admin/revenue", object="Doanh thu",
         steps="Click Doanh thu",
         expected="Trang thống kê doanh thu, lọc theo ngày, export báo cáo"),
    base(screen="Admin Dashboard", feature="Export revenue", component="GET /admin/revenue/export", object="Export báo cáo",
         steps="Click 'Xuất báo cáo'",
         expected="Download file xlsx/csv chứa doanh thu"),
    base(screen="Admin Dashboard", feature="Doanh thu chart", component="Include all status", object="Tính doanh thu",
         pre="Có order với status completed, shipping, delivered",
         steps="Quan sát chart",
         expected="Doanh thu cộng từ tất cả status được coi là 'thực thu'",
         note="Fix commit cb952ae"),

    base(screen="Admin Dashboard", feature="Order detail", component="Xem chi tiết", object="Click 1 đơn",
         steps="Click vào đơn",
         expected="Trang chi tiết đơn, mark viewed_at=now() ngay khi mở"),
    base(screen="Admin Dashboard", feature="Order status update", component="Update status", object="Cập nhật trạng thái",
         steps="PATCH /admin/orders/{order}/status",
         expected="Status đổi, log change"),
    base(screen="Admin Dashboard", feature="User management", component="Khoá user", object="Khoá/mở user",
         steps="Toggle is_locked",
         expected="users.is_locked đổi, user không login được"),
    base(screen="Admin Dashboard", feature="User management", component="Xoá user", object="Soft delete",
         steps="Click Xoá",
         expected="User bị xoá, không hiện trong danh sách"),

    base(screen="Admin Dashboard", feature="Auth", component="Yêu cầu admin guard", object="Truy cập khi không phải admin",
         steps="User thường vào /admin/dashboard",
         expected="Redirect /admin/login", type="Access Control & Security"),
    base(screen="Admin Dashboard", feature="CSRF", component="POST không token", object="CSRF",
         steps="POST /admin/users/{id} không _token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="Admin Dashboard", feature="Authz", component="User không thể vào admin route", object="Bypass",
         steps="User thường gọi DELETE /admin/users/{id}",
         expected="HTTP 403", type="Access Control & Security"),

    base(screen="Admin Dashboard", feature="Data Integrity", component="Tổng doanh thu = SUM(order.total)", object="Tính tổng",
         steps="So với SUM(orders.total_amount) tại các status doanh thu",
         expected="Trùng khớp", type="Data Integrity"),
    base(screen="Admin Dashboard", feature="Performance", component="Trang load lớn",
         object="DB có 10000 orders",
         steps="Mở /admin/dashboard",
         expected="< 2s render (truy vấn aggregate có index)", type="Performance Test"),

    base(screen="Admin Dashboard", feature="Browser",    component="Chrome",  object="OK", steps="Mở /admin/dashboard", expected="OK", type="Compatibility Test"),
    base(screen="Admin Dashboard", feature="Browser",    component="Firefox", object="OK", steps="Mở /admin/dashboard", expected="OK", type="Compatibility Test"),
    base(screen="Admin Dashboard", feature="Responsive", component="Tablet",  object="768px", steps="DevTools", expected="Sidebar collapse được", type="Compatibility Test"),
]


# ==================== ADMIN PRODUCTS / CATEGORIES ====================
ADMIN_PRODUCT_CASES = [
    base(screen="Admin Products", feature="List", component="Trang index", object="Trang /admin/products",
         pre="Đã login admin",
         steps="Truy cập /admin/products",
         expected="Bảng sản phẩm: ảnh, tên, danh mục, giá, stock, trạng thái, nút Sửa/Xoá, có Search & Filter", type="UI"),
    base(screen="Admin Products", feature="Search", component="Theo tên", object="Search",
         steps="Nhập từ khoá",
         expected="Filter đúng"),
    base(screen="Admin Products", feature="Filter", component="Theo danh mục", object="Filter category",
         steps="Chọn category dropdown",
         expected="List lọc đúng"),

    base(screen="Admin Products", feature="Create", component="Tạo mới", object="Tạo sản phẩm",
         steps="1. Truy cập /admin/products/create\n2. Điền form, upload ảnh\n3. Submit",
         expected="Tạo product, redirect list với flash success, ảnh lưu trong storage"),
    base(screen="Admin Products", feature="Edit",   component="Cập nhật",   object="Update sản phẩm",
         pre="Có product cần sửa",
         steps="PATCH /admin/products/{id}",
         expected="Cập nhật DB"),
    base(screen="Admin Products", feature="Delete", component="Xoá",       object="Xoá sản phẩm",
         pre="Có product",
         steps="DELETE /admin/products/{id}",
         expected="Xoá hoặc soft delete, redirect"),
    base(screen="Admin Products", feature="Toggle featured", component="is_featured", object="Đánh dấu nổi bật",
         steps="Toggle is_featured",
         expected="Sản phẩm xuất hiện trên trang chủ"),
    base(screen="Admin Products", feature="Image upload", component="Multi image", object="Upload nhiều ảnh",
         steps="Chọn nhiều file ảnh",
         expected="Tất cả ảnh upload, ảnh đầu là primary"),
    base(screen="Admin Products", feature="Variants", component="Sizes/Colors", object="Nhập sizes/colors",
         steps="Nhập sizes, colors, materials",
         expected="Lưu đúng JSON array"),
    base(screen="Admin Products", feature="Export", component="GET /admin/products/export", object="Export xlsx",
         steps="Click Export",
         expected="Download file chứa sản phẩm"),
    base(screen="Admin Products", feature="Import", component="POST /admin/products/import", object="Import xlsx",
         steps="Upload file mẫu",
         expected="Import record, báo cáo số ok/lỗi"),

    base(screen="Admin Products", feature="Validation", component="Tên rỗng", object="Bỏ tên",
         steps="Submit không name",
         expected="Báo 'Vui lòng nhập tên'", type="Abnormal"),
    base(screen="Admin Products", feature="Validation", component="Giá âm",  object="price=-1000",
         steps="Submit price=-1000",
         expected="Validate fail", type="Abnormal"),
    base(screen="Admin Products", feature="Validation", component="Stock âm", object="stock=-5",
         steps="Submit stock=-5",
         expected="Validate fail", type="Abnormal"),
    base(screen="Admin Products", feature="Validation", component="Ảnh sai loại", object="Upload .exe",
         steps="Upload file .exe",
         expected="Validate fail", type="Abnormal"),
    base(screen="Admin Products", feature="Validation", component="Slug trùng", object="Trùng slug",
         pre="Slug đã tồn tại",
         steps="Submit slug trùng",
         expected="Validate fail hoặc tự append số", type="Abnormal"),

    base(screen="Admin Categories", feature="List", component="Trang index", object="/admin/categories",
         steps="Truy cập",
         expected="Bảng danh mục - không còn cột Parent Category sau khi gỡ",
         note="Fix commit 365a5c9 - đã gỡ cột parent category"),
    base(screen="Admin Categories", feature="Create", component="Form", object="Form tạo danh mục",
         steps="Truy cập /admin/categories/create",
         expected="Form không còn dropdown 'Parent Category' (đã gỡ)",
         note="Fix commit 377c5b6"),
    base(screen="Admin Categories", feature="Edit", component="Form sửa", object="Sửa danh mục",
         steps="Truy cập /admin/categories/{id}/edit",
         expected="Form đúng spec hiện tại, không trường parent"),
    base(screen="Admin Categories", feature="Delete", component="Xoá danh mục có sản phẩm", object="Foreign key",
         pre="Category có sản phẩm",
         steps="Xoá",
         expected="Báo lỗi 'Không thể xoá danh mục đang có sản phẩm'", type="Abnormal"),

    base(screen="Admin Products", feature="DB", component="Slug unique", object="Slug duy nhất",
         steps="Đếm products.slug",
         expected="Không có 2 record trùng slug active", type="Data Integrity"),
    base(screen="Admin Products", feature="DB", component="image path", object="Path đúng prefix",
         pre="Vừa upload ảnh",
         steps="Đọc products.image",
         expected="Path bắt đầu bằng 'products/...' (storage)", type="Data Integrity"),

    base(screen="Admin Products", feature="Auth", component="Yêu cầu admin", object="User thường gọi route",
         steps="User thường POST /admin/products",
         expected="HTTP 403 / redirect /admin/login", type="Access Control & Security"),
    base(screen="Admin Products", feature="Security", component="Upload php", object="exploit.php",
         steps="Upload exploit.php",
         expected="Block do mime validate", type="Access Control & Security"),
    base(screen="Admin Products", feature="Security", component="XSS tên sản phẩm", object="Tên chứa <script>",
         steps="Tạo product với tên XSS",
         expected="Render escape trong tất cả nơi", type="Access Control & Security"),

    base(screen="Admin Products", feature="Performance", component="List 5000 sản phẩm", object="Phân trang lớn",
         steps="Mở /admin/products khi DB 5000 record",
         expected="< 1.5s render (paginate + index)", type="Performance Test"),
    base(screen="Admin Products", feature="Browser", component="Chrome",  object="OK", steps="Mở admin", expected="OK", type="Compatibility Test"),
    base(screen="Admin Products", feature="Browser", component="Firefox", object="OK", steps="Mở admin", expected="OK", type="Compatibility Test"),
]


# ==================== ADMIN - ORDERS ====================
ADMIN_ORDERS_CASES = [
    base(screen="Admin Orders", feature="Layout", component="Trang index", object="/admin/orders",
         pre="Đã login admin",
         steps="Truy cập /admin/orders",
         expected="Bảng đơn: mã, khách hàng, ngày, tổng, trạng thái, badge NEW, nút Xem", type="UI"),
    base(screen="Admin Orders", feature="Search", component="Ô tìm kiếm", object="Hiển thị ô search",
         steps="Quan sát",
         expected="Có ô search theo mã đơn / tên khách + dropdown lọc theo trạng thái", type="UI"),
    base(screen="Admin Orders", feature="Order detail", component="Trang show", object="/admin/orders/{order}",
         steps="Click vào 1 đơn",
         expected="Trang chi tiết: thông tin KH, ship address, list items, tổng tiền, status, lịch sử", type="UI"),
    base(screen="Admin Orders", feature="Status badge", component="Màu sắc", object="Hiển thị màu theo status",
         steps="Quan sát badge",
         expected="pending = vàng, confirmed = xanh dương, shipping = cam, completed = xanh lá, cancelled = đỏ", type="UI"),

    base(screen="Admin Orders", feature="Search", component="Theo mã đơn", object="Search mã",
         steps="Nhập 'ORD2026' vào ô search",
         expected="Lọc đúng order_code chứa từ khoá"),
    base(screen="Admin Orders", feature="Search", component="Theo tên khách", object="Search tên",
         steps="Nhập tên khách",
         expected="Lọc đúng (whereHas user)"),
    base(screen="Admin Orders", feature="Filter", component="Theo trạng thái", object="Lọc status",
         steps="Chọn dropdown status=shipping",
         expected="Chỉ hiện đơn shipping"),
    base(screen="Admin Orders", feature="Combo", component="Search + Filter", object="Kết hợp",
         steps="Search + Filter status",
         expected="Cả 2 điều kiện AND, query string giữ nguyên khi sang trang"),
    base(screen="Admin Orders", feature="Pagination", component="Phân trang", object="Page=2",
         pre="DB > 10 đơn",
         steps="Click trang 2",
         expected="Hiển thị batch tiếp theo, URL giữ ?search & ?status"),
    base(screen="Admin Orders", feature="NEW badge", component="3 đơn mới", object="Top 3 chưa xem",
         pre="DB có > 3 đơn chưa view",
         steps="Quan sát badge",
         expected="Tối đa 3 đơn mới nhất có badge NEW (theo is_new + viewed_at IS NULL)",
         note="Logic fix tại commit c48c56f"),
    base(screen="Admin Orders", feature="Mark viewed", component="Click show", object="Tự đánh dấu đã xem",
         pre="Order viewed_at NULL",
         steps="Mở /admin/orders/{order}",
         expected="orders.viewed_at = now(), badge NEW biến mất khi quay lại list"),
    base(screen="Admin Orders", feature="Status update", component="Pending → Confirmed", object="Xác nhận đơn",
         pre="Order status=pending",
         steps="PATCH /admin/orders/{order}/status status=confirmed",
         expected="DB cập nhật status=confirmed, lưu vào lịch sử"),
    base(screen="Admin Orders", feature="Status update", component="Confirmed → Shipping", object="Bắt đầu giao",
         pre="Order status=confirmed",
         steps="PATCH status=shipping",
         expected="DB cập nhật status=shipping"),
    base(screen="Admin Orders", feature="Status update", component="Shipping → Delivered", object="Đã giao",
         pre="Order status=shipping",
         steps="PATCH status=delivered/completed",
         expected="Doanh thu được tính, customer có thể đánh giá sản phẩm"),
    base(screen="Admin Orders", feature="Status update", component="Hủy đơn", object="Cancel",
         pre="Order status=pending hoặc confirmed",
         steps="PATCH status=cancelled",
         expected="Status=cancelled, hoàn lại stock (stock_deducted=0), không tính doanh thu"),

    base(screen="Admin Orders", feature="Status update", component="Cancelled lock", object="Đã hủy không đổi lại",
         pre="Order status=cancelled",
         steps="PATCH status=confirmed",
         expected="Báo lỗi 'Đơn đã hủy không thể chuyển lại trạng thái khác vì sẽ lệch tồn kho'", type="Abnormal"),
    base(screen="Admin Orders", feature="Validation", component="Status sai", object="Status không nằm trong enum",
         steps="PATCH status=hacked",
         expected="Validate fail (in:pending,confirmed,shipping,completed,cancelled,cod,paid,delivered)", type="Abnormal"),
    base(screen="Admin Orders", feature="Edge", component="Đơn tương lai", object="Đơn created_at > now",
         pre="DB có đơn created_at trong tương lai",
         steps="Mở /admin/orders",
         expected="Đơn tương lai bị xoá (auto cleanup ở index)",
         note="Logic làm sạch theo commit 93049b3 (rời future orders)", type="Abnormal"),
    base(screen="Admin Orders", feature="Stock", component="Hủy → hoàn kho", object="Stock hoàn về",
         pre="Order completed có sản phẩm trừ stock 5",
         steps="Đổi sang cancelled",
         expected="Stock + 5, stock_deducted=0", type="Data Integrity"),
    base(screen="Admin Orders", feature="Total amount", component="Match với items", object="Tổng tiền đúng",
         steps="So orders.total_amount với SUM(order_items)",
         expected="Trùng khớp (đã trừ voucher + cộng ship)", type="Data Integrity"),

    base(screen="Admin Orders", feature="Auth", component="Yêu cầu admin", object="User thường truy cập",
         steps="GET /admin/orders bằng user thường",
         expected="Redirect /admin/login", type="Access Control & Security"),
    base(screen="Admin Orders", feature="CSRF", component="PATCH không token", object="CSRF guard",
         steps="PATCH /admin/orders/{order}/status không _token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="Admin Orders", feature="Performance", component="Load 5k đơn", object="Phân trang",
         steps="Mở list khi DB có 5000 order",
         expected="< 1.5s (eager-load user, paginate 10)", type="Performance Test"),
    base(screen="Admin Orders", feature="Performance", component="N+1", object="Eager load user",
         steps="Bật Debugbar",
         expected="< 5 query / 1 page (đã with('user'))", type="Performance Test"),
    base(screen="Admin Orders", feature="Browser", component="Chrome",  object="OK", steps="Mở /admin/orders", expected="OK", type="Compatibility Test"),
    base(screen="Admin Orders", feature="Browser", component="Firefox", object="OK", steps="Mở /admin/orders", expected="OK", type="Compatibility Test"),
    base(screen="Admin Orders", feature="Responsive", component="Tablet 768px", object="Bảng cuộn ngang", steps="DevTools", expected="Table có scroll-x", type="Compatibility Test"),
]


# ==================== ADMIN - CUSTOMERS & ADMIN ACCOUNTS ====================
ADMIN_USERS_CASES = [
    base(screen="Admin Customers", feature="Layout", component="Trang index", object="/admin/users",
         pre="Đã login admin",
         steps="Truy cập /admin/users",
         expected="Bảng khách hàng: avatar, tên, email, sđt, ngày đăng ký, trạng thái, action", type="UI"),
    base(screen="Admin Customers", feature="Layout", component="Tab Admins", object="/admin/users/admins",
         steps="Click tab 'Tài khoản admin'",
         expected="Hiển thị danh sách admin trong table admins", type="UI"),
    base(screen="Admin Customers", feature="Detail", component="Trang show", object="/admin/users/{user}",
         steps="Click 'Xem' user",
         expected="Hiển thị thông tin chi tiết + lịch sử đơn hàng của user đó", type="UI"),

    base(screen="Admin Customers", feature="Search", component="Theo tên/email", object="Filter",
         steps="Nhập keyword",
         expected="Lọc đúng"),
    base(screen="Admin Customers", feature="Lock user", component="Khoá tài khoản", object="is_locked=1",
         pre="User chưa khoá",
         steps="Click 'Khoá' user",
         expected="users.is_locked=1, user không login được"),
    base(screen="Admin Customers", feature="Unlock user", component="Mở khoá", object="is_locked=0",
         pre="User đang khoá",
         steps="Click 'Mở khoá'",
         expected="users.is_locked=0, user login được lại"),
    base(screen="Admin Customers", feature="Delete user", component="Xoá customer", object="DELETE /admin/users/{user}",
         pre="User không có đơn liên kết",
         steps="Click xoá",
         expected="User bị xoá, không hiện trong list"),

    base(screen="Admin Customers", feature="Create admin", object="Tạo tài khoản admin", component="POST /admin/users",
         pre="Bảng admins tồn tại",
         steps="1. /admin/users/create\n2. Điền name, email, password\n3. Submit",
         expected="Tạo record trong bảng admins, password được hash"),
    base(screen="Admin Customers", feature="Validation", component="Tạo admin email trùng", object="Email đã tồn tại",
         pre="Email A đã có trong admins",
         steps="Submit create admin email=A",
         expected="Validate fail 'email phải unique'", type="Abnormal"),
    base(screen="Admin Customers", feature="Validation", component="Password ngắn", object="< 8 ký tự",
         steps="Submit password='123'",
         expected="Validate fail", type="Abnormal"),
    base(screen="Admin Customers", feature="Validation", component="Password confirm sai", object="Mismatch",
         steps="password != password_confirmation",
         expected="Validate fail", type="Abnormal"),

    base(screen="Admin Customers", feature="Edge", component="Xoá user có order", object="Foreign key",
         pre="User có order",
         steps="Xoá user",
         expected="Báo lỗi hoặc soft delete, không gãy DB", type="Abnormal"),
    base(screen="Admin Customers", feature="Edge", component="Tự khoá admin chính mình", object="Self lock",
         pre="Đang đăng nhập admin A",
         steps="Cố khoá chính tài khoản A",
         expected="Không cho phép, hoặc cảnh báo", type="Abnormal"),

    base(screen="Admin Customers", feature="DB", component="Hash password admin", object="Mật khẩu hash",
         pre="Vừa tạo admin",
         steps="Đọc admins.password trong DB",
         expected="bcrypt hash, không plain text", type="Data Integrity"),
    base(screen="Admin Customers", feature="DB", component="is_locked default", object="Mặc định = 0",
         pre="Vừa register user",
         steps="Đọc users.is_locked",
         expected="= 0 (false)", type="Data Integrity"),

    base(screen="Admin Customers", feature="Auth", component="Yêu cầu admin", object="User thường truy cập",
         steps="User truy cập /admin/users",
         expected="Redirect /admin/login", type="Access Control & Security"),
    base(screen="Admin Customers", feature="Authz", component="User không xoá admin", object="Inject DELETE",
         steps="User thường gửi DELETE /admin/users/{id}",
         expected="HTTP 403", type="Access Control & Security"),
    base(screen="Admin Customers", feature="CSRF", component="POST không token", object="Bảo vệ POST",
         steps="POST /admin/users không _token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="Admin Customers", feature="Security", component="XSS tên user", object="Render an toàn",
         steps="User có name=<script>",
         expected="Blade escape, không thực thi", type="Access Control & Security"),

    base(screen="Admin Customers", feature="Performance", component="List 10k users", object="Paginate",
         steps="Mở /admin/users với 10k record",
         expected="< 1.5s render", type="Performance Test"),
    base(screen="Admin Customers", feature="Browser", component="Chrome",  object="OK", steps="Mở /admin/users", expected="OK", type="Compatibility Test"),
    base(screen="Admin Customers", feature="Browser", component="Firefox", object="OK", steps="Mở /admin/users", expected="OK", type="Compatibility Test"),
    base(screen="Admin Customers", feature="Responsive", component="Tablet", object="768px", steps="DevTools", expected="Bảng cuộn ngang được", type="Compatibility Test"),
]


# ==================== ADMIN - VOUCHERS ====================
ADMIN_VOUCHER_CASES = [
    base(screen="Admin Vouchers", feature="Layout", component="Trang index", object="/admin/vouchers",
         pre="Đã login admin",
         steps="Truy cập /admin/vouchers",
         expected="Bảng voucher: code, name, type (percent/fixed), value, đk min, hạn dùng, trạng thái", type="UI"),
    base(screen="Admin Vouchers", feature="Create form", component="Trang create", object="/admin/vouchers/create",
         steps="Click 'Thêm voucher'",
         expected="Form: code, name, type, value, min_order_amount, max_discount_amount, usage_limit, starts_at, ends_at, is_active", type="UI"),
    base(screen="Admin Vouchers", feature="Edit form", component="Trang edit", object="/admin/vouchers/{voucher}/edit",
         steps="Click 'Sửa'",
         expected="Form đầy đủ field, pre-fill dữ liệu hiện có", type="UI"),

    base(screen="Admin Vouchers", feature="Search", component="Theo code/name", object="Filter list",
         steps="Search keyword",
         expected="Lọc đúng record"),
    base(screen="Admin Vouchers", feature="Create OK", component="Voucher percent", object="Tạo voucher %",
         steps="1. Code=SUMMER10\n2. type=percent, value=10\n3. ends_at 30 ngày sau\n4. Submit",
         expected="Voucher lưu DB với code='SUMMER10' (UPPERCASE), redirect index"),
    base(screen="Admin Vouchers", feature="Create OK", component="Voucher fixed", object="Tạo voucher tiền cố định",
         steps="type=fixed, value=50000",
         expected="Lưu DB type=fixed, value=50000"),
    base(screen="Admin Vouchers", feature="Update OK", component="Sửa voucher", object="Update voucher",
         steps="PATCH /admin/vouchers/{id}",
         expected="Cập nhật thành công"),
    base(screen="Admin Vouchers", feature="Delete", component="Xoá voucher", object="Xoá",
         steps="DELETE /admin/vouchers/{id}",
         expected="Xoá khỏi DB hoặc soft delete"),
    base(screen="Admin Vouchers", feature="Toggle active", component="is_active", object="Bật/tắt voucher",
         steps="Toggle checkbox is_active",
         expected="DB cập nhật, voucher tắt không apply được ở checkout"),

    base(screen="Admin Vouchers", feature="Validation", component="Code rỗng", object="Bỏ code",
         steps="Submit không code",
         expected="Báo 'Vui lòng nhập code'", type="Abnormal"),
    base(screen="Admin Vouchers", feature="Validation", component="Code trùng", object="Unique constraint",
         pre="Code 'X' đã tồn tại",
         steps="Tạo voucher code='X' lần nữa",
         expected="Validate fail 'Mã đã tồn tại'", type="Abnormal"),
    base(screen="Admin Vouchers", feature="Validation", component="Value <= 0", object="Giá trị invalid",
         steps="value=0 hoặc -10",
         expected="Validate fail (min:0.01)", type="Abnormal"),
    base(screen="Admin Vouchers", feature="Validation", component="Type không hợp lệ", object="type=abc",
         steps="Submit type='abc'",
         expected="Validate fail (in:percent,fixed)", type="Abnormal"),
    base(screen="Admin Vouchers", feature="Validation", component="ends_at trước starts_at", object="Ngày kết thúc sớm hơn bắt đầu",
         steps="starts_at=hôm nay, ends_at=hôm qua",
         expected="Validate fail 'ends_at after_or_equal starts_at'", type="Abnormal"),
    base(screen="Admin Vouchers", feature="Validation", component="usage_limit < 1", object="usage_limit=0",
         steps="Submit usage_limit=0",
         expected="Validate fail", type="Abnormal"),

    base(screen="Admin Vouchers", feature="DB", component="Code uppercase", object="Auto upper",
         pre="Submit code='summer10'",
         steps="Đọc vouchers.code",
         expected="DB lưu 'SUMMER10' (Str::upper)", type="Data Integrity"),
    base(screen="Admin Vouchers", feature="DB", component="usage_limit", object="Đếm số lần dùng",
         pre="Voucher có usage_limit=5",
         steps="Sau khi 5 user dùng",
         expected="Voucher tự disable / không apply được", type="Data Integrity"),
    base(screen="Admin Vouchers", feature="DB", component="Voucher hết hạn", object="Theo ends_at",
         pre="ends_at < now()",
         steps="User apply ở checkout",
         expected="Báo voucher hết hạn", type="Data Integrity"),

    base(screen="Admin Vouchers", feature="Auth", component="Yêu cầu admin", object="Guest truy cập",
         steps="GET /admin/vouchers chưa login admin",
         expected="Redirect /admin/login", type="Access Control & Security"),
    base(screen="Admin Vouchers", feature="CSRF", component="Không token", object="POST không token",
         steps="POST /admin/vouchers không _token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="Admin Vouchers", feature="Security", component="XSS name voucher", object="Render an toàn",
         steps="Tạo voucher name=<script>alert(1)</script>",
         expected="Blade escape", type="Access Control & Security"),

    base(screen="Admin Vouchers", feature="Performance", component="List 500 voucher", object="Tải nhanh",
         steps="Mở /admin/vouchers", expected="< 1s render", type="Performance Test"),
    base(screen="Admin Vouchers", feature="Browser", component="Chrome",  object="OK", steps="Mở", expected="OK", type="Compatibility Test"),
    base(screen="Admin Vouchers", feature="Browser", component="Firefox", object="OK", steps="Mở", expected="OK", type="Compatibility Test"),
    base(screen="Admin Vouchers", feature="Responsive", component="Tablet", object="768px", steps="DevTools", expected="OK", type="Compatibility Test"),
]


# ==================== ADMIN - REVIEWS & CONTACT MESSAGES ====================
ADMIN_REV_MSG_CASES = [
    # Reviews
    base(screen="Admin Reviews", feature="Layout", component="Trang index", object="/admin/reviews",
         pre="Đã login admin",
         steps="Truy cập /admin/reviews",
         expected="Bảng review: tên sản phẩm, user, rating, nội dung, ngày, trạng thái, nút Xoá", type="UI"),
    base(screen="Admin Reviews", feature="Filter", component="Theo rating", object="Lọc rating thấp",
         steps="Lọc rating=1",
         expected="Chỉ review 1 sao hiện", type="UI"),
    base(screen="Admin Reviews", feature="Delete", component="Xoá review", object="DELETE /admin/reviews/{id}",
         pre="Có review",
         steps="Click xoá",
         expected="Record bị xoá, average rating tính lại"),
    base(screen="Admin Reviews", feature="Sort", component="Mới nhất", object="Sort created_at desc",
         steps="Quan sát thứ tự",
         expected="Mặc định review mới nhất lên đầu"),

    base(screen="Admin Reviews", feature="Auth", component="Yêu cầu admin", object="Guest truy cập",
         steps="GET /admin/reviews chưa login",
         expected="Redirect login admin", type="Access Control & Security"),
    base(screen="Admin Reviews", feature="CSRF", component="DELETE không token", object="CSRF",
         steps="DELETE không token",
         expected="HTTP 419", type="Access Control & Security"),

    # Contact Messages
    base(screen="Admin Contact", feature="Layout", component="Trang index", object="/admin/contact-messages",
         pre="Đã login admin",
         steps="Truy cập",
         expected="Danh sách tin nhắn: name, email, subject, status (new/read/replied), ngày gửi", type="UI"),
    base(screen="Admin Contact", feature="Filter", component="Theo status", object="Lọc 'new'",
         steps="Chọn status=new",
         expected="Chỉ hiện tin nhắn chưa đọc"),
    base(screen="Admin Contact", feature="Search", component="Theo nội dung", object="Tìm trong message",
         steps="Search keyword",
         expected="Lọc đúng (name/email/subject/message)"),
    base(screen="Admin Contact", feature="Detail", component="Trang show", object="/admin/contact-messages/{msg}",
         steps="Click 'Xem'",
         expected="Hiển thị thông tin gửi, nội dung tin nhắn, danh sách phản hồi"),
    base(screen="Admin Contact", feature="Mark read", component="Tự đánh dấu", object="status: new → read",
         pre="Message status=new",
         steps="Mở /admin/contact-messages/{msg}",
         expected="read_at=now(), status đổi thành 'read'"),
    base(screen="Admin Contact", feature="Reply", component="Gửi phản hồi", object="POST reply",
         pre="Tin nhắn cần phản hồi",
         steps="1. Nhập subject, body\n2. Submit",
         expected="Tạo record contact_replies, gửi mail tới khách, status='replied', replied_at=now()"),
    base(screen="Admin Contact", feature="Reply history", component="Nhiều phản hồi", object="Lưu nhiều replies",
         pre="Đã có 2 reply",
         steps="Mở show",
         expected="Hiển thị tất cả replies theo thứ tự thời gian"),
    base(screen="Admin Contact", feature="Delete", component="Xoá tin nhắn", object="DELETE /admin/contact-messages/{msg}",
         steps="Click xoá",
         expected="Tin nhắn + replies cascade xoá"),

    base(screen="Admin Contact", feature="Validation", component="Reply rỗng", object="Bỏ body",
         steps="Submit reply không body",
         expected="Validate fail 'Vui lòng nhập nội dung'", type="Abnormal"),
    base(screen="Admin Contact", feature="Validation", component="Body quá dài", object="> 3000 ký tự",
         steps="Submit 5000 ký tự",
         expected="Validate fail", type="Abnormal"),
    base(screen="Admin Contact", feature="Edge", component="Mail SMTP fail", object="SMTP down",
         pre="SMTP cấu hình sai",
         steps="Click gửi reply",
         expected="Hiển thị warning 'Mail không gửi được', record vẫn lưu nhưng sent_at=NULL",
         note="Có try/catch quanh Mail::send", type="Abnormal"),

    base(screen="Admin Contact", feature="DB", component="contact_replies.admin_id", object="Lưu ID admin gửi",
         pre="Admin A reply",
         steps="Check contact_replies.admin_id",
         expected="= admin A id", type="Data Integrity"),
    base(screen="Admin Contact", feature="DB", component="replied_at", object="Cập nhật thời gian",
         pre="Reply thành công",
         steps="Check contact_messages.replied_at",
         expected="= timestamp khi reply", type="Data Integrity"),

    base(screen="Admin Contact", feature="Auth", component="Yêu cầu admin", object="Guest truy cập",
         steps="GET /admin/contact-messages chưa login",
         expected="Redirect /admin/login", type="Access Control & Security"),
    base(screen="Admin Contact", feature="CSRF", component="Reply không token", object="CSRF",
         steps="POST reply không _token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="Admin Contact", feature="Security", component="XSS body reply", object="Render an toàn trong mail",
         steps="Reply body chứa <script>",
         expected="Mail HTML escape, không thực thi", type="Access Control & Security"),

    base(screen="Admin Contact", feature="Performance", component="List 2000 msg", object="Phân trang",
         steps="Mở list", expected="< 1.5s", type="Performance Test"),
    base(screen="Admin Contact", feature="Browser", component="Chrome",  object="OK", steps="Mở", expected="OK", type="Compatibility Test"),
    base(screen="Admin Contact", feature="Responsive", component="Tablet", object="768px", steps="DevTools", expected="OK", type="Compatibility Test"),
]


# ==================== ADMIN - SETTINGS & PROFILE ====================
ADMIN_SETTING_CASES = [
    # Website Settings
    base(screen="Admin Settings", feature="Layout", component="Trang edit", object="/admin/settings",
         pre="Đã login admin",
         steps="Truy cập /admin/settings",
         expected="Form đầy đủ: site_name, tagline, support_email, hotline, address, hero, meta, social, logo, hero_image", type="UI"),
    base(screen="Admin Settings", feature="Logo upload", component="File input", object="Upload logo",
         steps="Chọn file logo .png, submit",
         expected="logo lưu vào storage/settings, hiển thị trên header sau khi cập nhật"),
    base(screen="Admin Settings", feature="Hero image", component="Upload", object="Upload hero",
         steps="Chọn ảnh hero, submit",
         expected="hero_image lưu, hiển thị trên trang chủ"),
    base(screen="Admin Settings", feature="Social links", component="Update facebook url", object="URL hợp lệ",
         steps="Nhập facebook_url='https://facebook.com/huonghoaxinh'",
         expected="Lưu DB, hiển thị icon FB trên footer link đúng"),
    base(screen="Admin Settings", feature="Meta SEO", component="Update meta", object="meta_title, description, keywords",
         steps="Nhập và submit",
         expected="<title> + <meta> trên trang chủ thay đổi"),
    base(screen="Admin Settings", feature="Featured limit", component="Số sản phẩm nổi bật", object="Đổi limit",
         steps="featured_products_limit=12, submit",
         expected="Trang chủ hiển thị tối đa 12 sản phẩm featured"),
    base(screen="Admin Settings", feature="Toggle", component="enable_reviews", object="Tắt review",
         steps="Bỏ tick enable_reviews",
         expected="Trang chi tiết sản phẩm ẩn section đánh giá"),
    base(screen="Admin Settings", feature="Catalog mode", component="enable_catalog_mode", object="Bật catalog mode",
         steps="Tick enable_catalog_mode",
         expected="Ẩn nút 'Thêm vào giỏ' & 'Mua ngay', chỉ hiển thị thông tin"),

    base(screen="Admin Settings", feature="Validation", component="site_name rỗng", object="Bỏ site_name",
         steps="Submit không site_name",
         expected="Validate fail 'Vui lòng nhập tên site'", type="Abnormal"),
    base(screen="Admin Settings", feature="Validation", component="Email sai format", object="support_email='abc'",
         steps="Submit",
         expected="Validate fail 'Email không hợp lệ'", type="Abnormal"),
    base(screen="Admin Settings", feature="Validation", component="URL FB sai", object="facebook_url='abc'",
         steps="Submit",
         expected="Validate fail 'URL không hợp lệ'", type="Abnormal"),
    base(screen="Admin Settings", feature="Validation", component="Logo quá lớn", object="> 2MB",
         steps="Upload logo 3MB",
         expected="Validate fail (max:2048)", type="Abnormal"),
    base(screen="Admin Settings", feature="Validation", component="featured_limit ngoài khoảng", object="< 4 hoặc > 24",
         steps="Submit featured_products_limit=2",
         expected="Validate fail (min:4 max:24)", type="Abnormal"),

    base(screen="Admin Settings", feature="DB", component="Lưu dạng key-value", object="WebsiteSetting table",
         pre="Vừa update",
         steps="Đọc website_settings table",
         expected="Mỗi setting là 1 row (key, value)", type="Data Integrity"),
    base(screen="Admin Settings", feature="Cache", component="Cache invalidate", object="Đổi setting → clear cache",
         pre="Cache settings đã warm",
         steps="Update setting",
         expected="Cache::forget được gọi, lần load tiếp lấy giá trị mới", type="Data Integrity"),

    # Admin Profile
    base(screen="Admin Profile", feature="Layout", component="Trang edit", object="/admin/profile",
         pre="Đã login admin",
         steps="Truy cập /admin/profile",
         expected="Form: name, email, phone, password (optional)", type="UI"),
    base(screen="Admin Profile", feature="Update", component="Sửa thông tin", object="Update profile",
         steps="Sửa name, submit",
         expected="DB cập nhật, flash success"),
    base(screen="Admin Profile", feature="Change password", component="Đổi mật khẩu admin", object="Đổi password",
         steps="Nhập current_password + new_password",
         expected="Hash mới lưu DB, login lại OK"),
    base(screen="Admin Profile", feature="Validation", component="Email trùng admin khác", object="Email đã có",
         steps="Đổi email = email admin khác",
         expected="Validate fail", type="Abnormal"),
    base(screen="Admin Profile", feature="Validation", component="Password mới sai confirm", object="Mismatch",
         steps="new != new_confirmation",
         expected="Validate fail", type="Abnormal"),

    base(screen="Admin Profile", feature="DB", component="Hash password mới", object="Lưu hash",
         pre="Vừa đổi pass",
         steps="Đọc admins.password",
         expected="bcrypt hash mới", type="Data Integrity"),

    base(screen="Admin Profile", feature="Auth", component="Guest truy cập", object="Yêu cầu login admin",
         steps="GET /admin/profile chưa login",
         expected="Redirect /admin/login", type="Access Control & Security"),
    base(screen="Admin Profile", feature="CSRF", component="PUT không token", object="CSRF",
         steps="PUT /admin/profile không _token",
         expected="HTTP 419", type="Access Control & Security"),
    base(screen="Admin Settings", feature="Security", component="Upload ảnh độc hại", object="exploit.php.png",
         steps="Upload file exploit",
         expected="Validate mime image, block file độc", type="Access Control & Security"),

    base(screen="Admin Settings", feature="Browser", component="Chrome",  object="OK", steps="Mở /admin/settings", expected="OK", type="Compatibility Test"),
    base(screen="Admin Settings", feature="Browser", component="Firefox", object="OK", steps="Mở /admin/settings", expected="OK", type="Compatibility Test"),
    base(screen="Admin Profile", feature="Responsive", component="Tablet", object="768px", steps="DevTools", expected="Form 1 cột trên tablet", type="Compatibility Test"),
]


# ---------------------------------------------------------------- Overview
def build_overview(wb, sheets_meta):
    """sheets_meta: list of dicts {name, total, last_row}."""
    ws = wb.create_sheet("Overview", 2)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [4, 26, 38, 10, 10, 10, 10, 10, 10, 12, 12, 12])

    ws["A1"] = "Overview - Test Case Coverage"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "System"
    ws["B2"] = SYSTEM_NAME
    ws["D2"] = "Created Date"
    ws["E2"] = DATE_TODAY
    ws["A3"] = "Author"
    ws["B3"] = TESTER
    for cell in ["A2","A3","D2"]:
        ws[cell].font = LABEL_FONT

    ws["A5"] = "1. Status Summary"
    ws["A5"].font = SECTION_FONT

    headers = ["#", "Sheet name", "Description", "Passed", "Failed",
               "Pending", "N/A", "Remain", "Total"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    style_header_row(ws, 6, len(headers), height=28)

    start_row = 7
    for i, meta in enumerate(sheets_meta):
        r = start_row + i
        sn = meta["name"]; total = meta["total"]; last_row = meta["last_row"]
        rng = f"'{sn}'!J11:J{last_row}"
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=2, value=sn)
        ws.cell(row=r, column=3, value=meta["desc"])
        ws.cell(row=r, column=4, value=f'=COUNTIF({rng},"Passed")')
        ws.cell(row=r, column=5, value=f'=COUNTIF({rng},"Failed")')
        ws.cell(row=r, column=6, value=f'=COUNTIF({rng},"Pending")')
        ws.cell(row=r, column=7, value=f'=COUNTIF({rng},"N/A")')
        ws.cell(row=r, column=8, value=f"=I{r}-D{r}-E{r}-F{r}-G{r}")
        ws.cell(row=r, column=9, value=total)
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                horizontal="center" if c != 2 and c != 3 else "left",
                indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 22

    total_row = start_row + len(sheets_meta)
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    for c in range(4, 10):
        col = get_column_letter(c)
        ws.cell(row=total_row, column=c,
                value=f"=SUM({col}{start_row}:{col}{total_row-1})")
    for c in range(1, 10):
        cell = ws.cell(row=total_row, column=c)
        cell.font = Font(bold=True)
        cell.fill = LABEL_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    # Test Type breakdown
    tt_start = total_row + 3
    ws.cell(row=tt_start, column=1, value="2. Test Type Breakdown").font = SECTION_FONT
    tt_headers = ["#", "Sheet name", "UI", "Normal", "Abnormal",
                  "Data Integrity", "Access Control & Security",
                  "Performance Test", "Compatibility Test", "Total"]
    for c, h in enumerate(tt_headers, start=1):
        ws.cell(row=tt_start+1, column=c, value=h)
    style_header_row(ws, tt_start+1, len(tt_headers), height=42)

    for i, meta in enumerate(sheets_meta):
        r = tt_start + 2 + i
        sn = meta["name"]; total = meta["total"]; last_row = meta["last_row"]
        rng = f"'{sn}'!I11:I{last_row}"
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=2, value=sn)
        for c_idx, t in enumerate(["UI","Normal","Abnormal","Data Integrity",
                                   "Access Control & Security",
                                   "Performance Test","Compatibility Test"], start=3):
            ws.cell(row=r, column=c_idx, value=f'=COUNTIF({rng},"{t}")')
        ws.cell(row=r, column=10, value=total)
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                horizontal="center" if c != 2 else "left", indent=1)
        ws.row_dimensions[r].height = 22

    final_row = tt_start + 2 + len(sheets_meta)
    ws.cell(row=final_row, column=1, value="Total")
    for c in range(3, 11):
        col = get_column_letter(c)
        ws.cell(row=final_row, column=c,
                value=f"=SUM({col}{tt_start+2}:{col}{final_row-1})")
    for c in range(1, 11):
        cell = ws.cell(row=final_row, column=c)
        cell.font = Font(bold=True)
        cell.fill = LABEL_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    # Legend
    legend_row = final_row + 3
    ws.cell(row=legend_row, column=1, value="3. Test Type Legend").font = SECTION_FONT
    legend = [
        ("UI",                          "Kiểm tra giao diện, layout, hiển thị"),
        ("Normal",                      "Luồng đúng (happy path), dữ liệu hợp lệ"),
        ("Abnormal",                    "Lỗi, dữ liệu không hợp lệ, edge case"),
        ("Data Integrity",              "Nhất quán dữ liệu so với DB, format, persistence"),
        ("Access Control & Security",   "Phân quyền, auth, SQLi/XSS/CSRF"),
        ("Performance Test",            "Hiệu năng, tốc độ phản hồi, race condition"),
        ("Compatibility Test",          "Tương thích trình duyệt + responsive"),
    ]
    for i, (k, v) in enumerate(legend):
        r = legend_row + 1 + i
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=r, column=3, value=v).border = BORDER
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", indent=1)


# ---------------------------------------------------------------- Main
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cover(wb)
    build_history(wb)

    sheets = [
        ("Home Screen",         "Home Screen",        "Trang chủ - banner, nổi bật, footer", HOME_CASES),
        ("Shop & Filter",       "Shop & Filter",      "Cửa hàng, tìm kiếm và lọc sản phẩm",  SHOP_CASES),
        ("Product Detail",      "Product Detail",     "Chi tiết sản phẩm",                   PRODUCT_CASES),
        ("Cart",                "Cart",               "Giỏ hàng",                            CART_CASES),
        ("Checkout & Payment",  "Checkout & Payment", "Thanh toán đơn hàng",                 CHECKOUT_CASES),
        ("Login",               "Login",              "Đăng nhập (customer & admin)",        LOGIN_CASES),
        ("Register",            "Register",           "Đăng ký tài khoản",                   REGISTER_CASES),
        ("My Account",          "My Account",         "Tài khoản cá nhân, lịch sử đơn",      ACCOUNT_CASES),
        ("Wishlist & Review",   "Wishlist & Review",  "Yêu thích & Đánh giá sản phẩm",       WISHLIST_CASES),
        ("Admin - Dashboard",   "Admin - Dashboard",  "Login admin, dashboard, doanh thu",   ADMIN_DASH_CASES),
        ("Admin - Products",    "Admin - Products",   "Quản lý sản phẩm & danh mục",         ADMIN_PRODUCT_CASES),
        ("Admin - Orders",      "Admin - Orders",     "Quản lý đơn hàng & trạng thái",       ADMIN_ORDERS_CASES),
        ("Admin - Customers",   "Admin - Customers",  "Quản lý khách hàng & admin accounts", ADMIN_USERS_CASES),
        ("Admin - Vouchers",    "Admin - Vouchers",   "Quản lý mã giảm giá",                 ADMIN_VOUCHER_CASES),
        ("Admin - Reviews & Contact","Admin - Reviews & Contact","Đánh giá & Tin nhắn liên hệ", ADMIN_REV_MSG_CASES),
        ("Admin - Settings & Profile","Admin - Settings & Profile","Cài đặt website & profile admin", ADMIN_SETTING_CASES),
    ]

    sheets_meta = []
    for sheet_name, title, desc, cases in sheets:
        build_detail_sheet(wb, sheet_name, title, cases)
        sheets_meta.append({
            "name": sheet_name, "desc": desc,
            "total": len(cases), "last_row": 10 + len(cases),
        })

    build_overview(wb, sheets_meta)

    out_path = r"C:\Users\Kieu Anh\Desktop\CD2\TC_HuongHoaXinh.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"\nTotal test cases: {sum(m['total'] for m in sheets_meta)}")
    for m in sheets_meta:
        print(f"  - {m['name']}: {m['total']} cases")

if __name__ == "__main__":
    main()
